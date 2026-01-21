import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
import re
import math
# win32com is Windows-only; Cloud Run runs Linux
try:
    import win32com.client as win32
    import win32.gencache
    _HAS_WIN32COM = True
except Exception:
    win32 = None
    _HAS_WIN32COM = False
import shutil
import sys

def clear_com_cache():
    if not _HAS_WIN32COM:
        return
    gen_py_path = os.path.join(os.environ['LOCALAPPDATA'], 'Temp', 'gen_py')
    if os.path.exists(gen_py_path):
        shutil.rmtree(gen_py_path)
    win32.gencache.is_readonly = False
    win32.gencache.Rebuild()

def set_dynamic_row_heights(ws, cover_row_map):
    
    for cover, rows in cover_row_map.items():
        first_row = rows[0]
        num_rows = len(rows)

        # Get text in column A of the first row
        cell_value = str(ws[f"A{first_row}"].value or "")
        num_chars = len(cell_value)

        # Calculate row height needed based on characters
        chars_per_line = 45
        px_per_line = 17
        calculated_height = ((num_chars - 1) // chars_per_line + 1) * px_per_line

        # Minimum height is 17 * number of rows
        min_required_height = px_per_line * num_rows

        # Use the max of calculated height and minimum height
        final_height = max(calculated_height, min_required_height)

        # Distribute height equally across rows
        height_per_row = final_height / num_rows

        # Set height for each row
        for row in rows:
            ws.row_dimensions[row].height = height_per_row


def map_covers_to_variables(covers_str, mapping):
    # Remove square brackets and split by ;
    covers = str(covers_str).strip('[]').split(';')
    # Map each cover to its variable, default to 'NA' if not found
    variables = [str(mapping.get(cover.strip(), 'NA')) for cover in covers]
    # Join back with ; and add brackets
    return '[' + ';'.join(variables) + ']'

def split_by_definition(df):
    # Extract all unique variables from x_def column
    all_variables = set()
    for def_val in df['x_def'].dropna():
        # Remove brackets and split by semicolon
        defs_str = str(def_val).strip('[]')
        vars_list = defs_str.split(';')
        for v in vars_list:
            if v != 'NA' and '&' in v:
                all_variables.update(v.split('&'))
            elif v != 'NA':
                all_variables.add(v)
    
    # Create new columns for each variable
    for var in all_variables:
        df[var] = None
    
    # Process each row
    for idx, row in df.iterrows():
        if pd.isna(row['x_def']) or pd.isna(row['x_val']):
            continue
            
        # Parse x_def and x_val
        def_str = str(row['x_def']).strip('[]')
        val_str = str(row['x_val']).strip('[]')
        def_values = def_str.split(';')
        val_values = val_str.split(';')
        
        # Create dictionaries to collect values for each variable
        var_values = {var: ['NA'] * len(def_values) for var in all_variables}
        
        # Map values to variables
        for i, (def_val, val) in enumerate(zip(def_values, val_values)):
            if def_val == 'NA' or val == 'NA':
                continue
                
            # Handle compound variables (with &)
            if '&' in def_val and '&' in val:
                sub_defs = def_val.split('&')
                sub_vals = val.split('&')
                
                for sub_def, sub_val in zip(sub_defs, sub_vals):
                    if sub_def != 'NA' and sub_val != 'NA':
                        var_values[sub_def][i] = sub_val
            else:
                var_values[def_val][i] = val
        
        # Update the dataframe with the collected values
        for var, values in var_values.items():
            df.at[idx, var] = '[' + ';'.join(values) + ']'
    # df.drop(columns=['x_val','x_def'],inplace=True)
    return df

def smart_convert(value):
    # Try to parse strings to numbers
    if isinstance(value, str):
        try:
            num = float(value)
            if num.is_integer():
                return int(num)
            return num
        except ValueError:
            return value

    # Convert floats with no decimals to int
    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value

def format_date(date_input):
    """Converts '01-04-2026' or Timestamp to '1st April 2026'"""
    if isinstance(date_input, pd.Timestamp):
        dt = date_input
    else:
        dt = datetime.strptime(date_input, "%d-%m-%Y")

    day = dt.day
    suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    return f"{day}{suffix} {dt.strftime('%B %Y')}"

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def parse_strike_counts(strikes):
    """Parse strike field to return tuple of strike counts per cover."""
    try:
        covers = strikes.strip("[]").split(";")
        return tuple(len(c.split("&")) if c != 'NA' else 0 for c in covers)
    except:
        return ()

def clean_column_names(df):
    rename_dict = {}
    for col in df.columns:
        new_col = col

        # Remove 'Var' (case-insensitive)
        new_col = re.sub(r'(?i)var', '', new_col).strip()

        # Cover X Basepay Wt -> Cover X Base Payout (case-insensitive, flexible spaces)
        new_col = re.sub(r'(Cover\s*\d+)__basepay_wt', r'\1 Base Payout', new_col, flags=re.IGNORECASE)

        # Cover X Strike SI Weights Y -> Cover X Strike Y SI Weight (case-insensitive)
        new_col = re.sub(r'(Cover\s*\d+)_Strike_SI_Weights_(\d+)', r'\1 Strike \2 SI Weight', new_col, flags=re.IGNORECASE)

        # Cover X_strike Y or Cover X Strike Y (underscore or space) -> Cover X Strike Y (normalize)
        new_col = re.sub(r'(Cover\s*\d+)_Strikes_?(\d+)', r'\1 Strike \2', new_col, flags=re.IGNORECASE)

        # Cover X_n_days -> Cover X No. of Days
        new_col = re.sub(r'(Cover\s*\d+)_n_days', r'\1 No. of Days', new_col, flags=re.IGNORECASE)

        # Cover X_deductible or Cover X Deductible -> Cover X Entry Trigger
        new_col = re.sub(r'(Cover\s*\d+)[ _]*deductible', r'\1 Entry Trigger', new_col, flags=re.IGNORECASE)

        # Cover X Ref_Lat_Lon -> Cover X Reference Lat Long
        new_col = re.sub(r'(Cover\s*\d+)[ _]*Ref_Lat_Lon', r'\1 Reference Lat Long', new_col, flags=re.IGNORECASE)

        # Specific direct mappings (case-insensitive)
        if new_col.lower() == 'loc_name':
            new_col = 'Location'
        elif new_col == 'RSD':
            new_col = 'Risk Start Date'
        elif new_col == 'RED':
            new_col = 'Risk End Date'
        elif new_col == 'Unit_Type':
            new_col = 'Unit Type'
        elif new_col == 'Unit_SI':
            new_col = 'Sum Insured'
        elif new_col == 'Unit_Net_Prem':
            new_col = 'Net Premium'
        elif new_col == 'Num_Risk_Units':
            new_col = 'No. of Risk Units'
        elif new_col == 'Deductible':
            new_col = 'Entry Trigger'
        elif new_col == 'Strikes':
            new_col = 'Strike'

        # General: replace underscores with spaces
        new_col = new_col.replace('_', ' ')

        # Capitalize each word (except all-uppercase abbreviations)
        new_col = ' '.join(word if word.isupper() else word.capitalize() for word in new_col.split())

        if new_col != col:
            rename_dict[col] = new_col

    return df.rename(columns=rename_dict)

def clear_all_borders(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border()

def apply_outer_border(sheet, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            
            # Get the existing borders
            current_border = cell.border
            
            # Combine existing borders with new ones
            cell.border = Border(
                left=thin_border.left if col == min_col else current_border.left,
                right=thin_border.right if col == max_col else current_border.right,
                top=thin_border.top if row == min_row else current_border.top,
                bottom=thin_border.bottom if row == max_row else current_border.bottom
            )
def insert_smart_page_breaks(ws,max_page_height=780):
    current_height = 0
    last_row = ws.UsedRange.Rows.Count
    row = 2  # Start after title/header rows

    while row <= last_row:
        # Skip blank rows (assumed to be non-table rows)
        while row <= last_row and all(ws.Cells(row, col).Value in [None, ""] for col in range(1, 7)):
            current_height += ws.Rows(row).RowHeight
            row += 1

        if row > last_row:
            break

        table_start = row
        table_height = 0

        # Measure the table’s height
        while row <= last_row and any(ws.Cells(row, col).Value not in [None, ""] for col in range(1, 7)):
            table_height += ws.Rows(row).RowHeight
            row += 1

        # Check if table fits in remaining space
        if current_height + table_height > max_page_height:
            ws.Rows(table_start).PageBreak = 1
            current_height = table_height
        else:
            current_height += table_height

def save_excel_sheet_as_pdf(excel_path, sheet_name, pdf_path, logo_path, formatted_time):
    clear_com_cache()
    if not _HAS_WIN32COM:
        print('PDF export skipped: win32com not available (Linux environment).')
        return
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False

    wb = excel.Workbooks.Open(os.path.abspath(excel_path))
    ws = wb.Sheets(sheet_name)

    # Set the print area (columns A to F, full used row range)
    last_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"A1:F{last_row}"

    # Page layout: Fit columns on one page, allow rows to go to multiple pages
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False  # Let rows span multiple pages

    ws.PageSetup.RightHeader = f"Termsheet ID - {formatted_time}"

    # Footer with "Powered by" and image
    ws.PageSetup.RightFooterPicture.Filename = os.path.abspath(logo_path)
    ws.PageSetup.RightFooter = 'Powered by &G'
    pic = ws.PageSetup.RightFooterPicture
    pic.Height = 25  # you can tweak this further
    pic.Width = 55   # reduce to your desired scale
    insert_smart_page_breaks(ws)
    # Export as PDF
    ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path))

    wb.Close(SaveChanges=False)
    excel.Quit()

def merge_c_to_f(ws, rows, start, end):
    for row in rows:
        start_col = start  # Column
        end_col = end    # Column
        cell_range = f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
        ws.merge_cells(cell_range)

def merge_center(ws,row_from, row_to):
    ws.merge_cells(str(row_from) + ':' + str(row_to))
    ws[row_from].alignment = Alignment(horizontal='center', vertical='center')


def parse_list(cell):
    """Strip [ ] and split on ';' (return empty list if NaN)."""
    if pd.isna(cell):
        return []
    s = str(cell).strip().lstrip('[').rstrip(']')
    return s.split(';') if s else []

def parse_cyclone_entries(cell):
    """From '(60,10,0.2)&(60,30,0.15)…' produce list of (speed,dist,weight)."""
    parts = cell.split('&')
    entries = []
    for p in parts:
        # each p is like '(60,10,0.2)'
        nums = p.strip('()').split(',')
        if len(nums)==3:
            speed, dist, wt = nums
            entries.append((float(speed), float(dist), float(wt)))
    return entries

def format_risk_location(loc_name):
    loc_fields = loc_name.strip('[]').split(';')
    lat, lon = loc_fields[0].split('_')

def format_risk_location(loc_name):
    loc_fields = loc_name.strip('[]').split(';')
    components = []

    if loc_fields[0] != "NA":
        lat, lon = loc_fields[0].split('_')
        components.append(f"Geo-Coordinates: {lat}°N {lon}°E")

    if loc_fields[1] != "NA":
        components.append(f"Pincode: {loc_fields[1]}")
    if loc_fields[2] != "NA":
        components.append(f"Block: {loc_fields[2]}")
    if loc_fields[3] != "NA":
        components.append(f"District: {loc_fields[3]}")
    if loc_fields[4] != "NA":
        components.append(f"State: {loc_fields[4]}")
    if len(loc_fields) > 5 and loc_fields[5] != "NA":
        components.append(f"Country: {loc_fields[5]}")
    else:
        components.append("Country: India")

    formatted_string = "\n".join(components)
    return formatted_string, len(components)

def convert_termsheet_uploader_to_draft(input_file):
    df_uploader = pd.read_excel(
        input_file,
        sheet_name='ts_uploader',
        na_values=[],           # Do not treat any string as NaN
        keep_default_na=False   # Prevent default NA-like strings ("NA", "NaN", etc.) from being parsed as NaN
    )    
    masters_df = pd.read_excel(input_file, sheet_name='Masters')
    split_index = masters_df[masters_df.iloc[:, 0].isna()].index[0]
    # Split into two tables
    first_table_df = masters_df.iloc[:split_index]
    second_table_df = masters_df.iloc[split_index + 1:].reset_index(drop=True)
    second_table_df.columns = second_table_df.iloc[0]
    second_table_df = second_table_df.loc[:, second_table_df.columns.notna()].reset_index(drop=True)
    cover_to_variable = dict(zip(second_table_df['Cover'], second_table_df['Variable']))
    df_uploader['x_def'] = df_uploader['Covers'].apply(lambda x: map_covers_to_variables(x, cover_to_variable))
    df_uploader = split_by_definition(df_uploader)
    bold_font = Font(bold=True)
    if pdf == "Yes":
        for location in range(df_uploader.shape[0]):
            now = datetime.now()
            formatted_time = now.strftime("%Y-%m-%d_%H-%M-%S") + f"-{now.microsecond:06d}"
            output_file = "Draft Termsheet Output"+ " "+ str(formatted_time) + ".xlsx"
            output_pdf = "Draft Termsheet Output"+ " " + str(formatted_time) + ".pdf"
            # Read data
            df = df_uploader.copy()
            # Convert each to dictionary
            masters_dict = first_table_df.set_index(first_table_df.columns[0]).to_dict(orient='index')
            covers_dict = second_table_df.set_index(second_table_df.columns[0]).to_dict(orient='index')
            row = df_uploader.iloc[location]
            unit_type = row['Unit_Type'].upper()
            risk_location, line_count = format_risk_location(row['loc_name'])
            risk_start = format_date(pd.to_datetime(row['RSD'], format="%d-%m-%Y"))
            risk_end = format_date(pd.to_datetime(row['RED'], format="%d-%m-%Y"))
            net_prem = float(row['Unit_Net_Prem'])
            gst = net_prem * float(row['GST(%)'])
            gross = net_prem + gst
            n_units = int(row['Num_Risk_Units'])
            unit_si = float(row['Unit_SI'])
            formatted_time = row['termsheet_id']
            # parse all the list‐columns
            ref_lat_long = parse_list(row['Ref_Lat_Lon'])
            covers      = parse_list(row['Covers'])
            sources     = parse_list(row['Data_Sources'])
            phases      = parse_list(row['Phase_Dates'])
            si_weights  = parse_list(row['SI_Weights'])
            strikes     = parse_list(row['Strikes'])
            strike_wts  = parse_list(row['Strike_SI_Weights'])
            var_deds    = parse_list(row['VAR_deductible'])
            var_lvls    = parse_list(row['VAR_levels'])
            var_szs     = parse_list(row['VAR_level_size'])
            var_bpays   = parse_list(row['VAR_basepay_wt'])
            var_types   = parse_list(row['VAR_payout_type'])
            cycl_payout = parse_list(row['Cyclone_Speed_Dist_PayoutWt'])

            # Build data as per the sample
            draft_data = [
                ['<<Name of Product>> (UIN: )', None, None, None, None, None],           # 1
                [None, None, None, None, None, None],                                    # 2
                ['RISK DETAILS', None, row['Risk_Details'], None, None, None],           # 3
                ['RISK LOCATION', None, None, None, None, None],                # 4
                [None, None, None, None, None, None],                                    # 5
                [None, None, None, None, None, None],                                    # 6
                [f'PREMIUM & COVERAGE SUMMARY (PER {unit_type})', None, None, None, None, None], # 7
                [None, None, None, None, None, None],                                    # 8
                ['RISK START DATE', None, risk_start, None, None, None],                 # 9
                ['RISK END DATE', None, risk_end, None, None, None],                     # 10
                [None, None, None, None, None, None],                                    # 11
                ['SUM INSURED', None, row['Unit_SI'], None, None, None],                 # 12
                ['PREMIUM AMOUNT', None, round(net_prem,4), None, None, None],                    # 13
                ['GST', None, round(gst,4), None, None, None],                              # 14
                ['GROSS PREMIUM', None, round(gross,4), None, None, None],                        # 15
                [None, None, None, None, None, None],                                    # 16
            ]
            columns = [
                '<<Name of Insurance Company>>', 'Unnamed: 1', 'Unnamed: 2',
                'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 5'
            ]
            df_draft = pd.DataFrame(draft_data, columns=columns)
            df_draft.to_excel(output_file, sheet_name='Draft Termsheet', index=False)

            # Now apply formatting with openpyxl
            wb = load_workbook(output_file)
            ws = wb['Draft Termsheet']
            clear_all_borders(ws)
            current_row = 19
            apply_outer_border(ws, 1,2, 1, 6)
            apply_outer_border(ws, 3,6, 1, 6)
            apply_outer_border(ws, 7,17, 1, 6)
            apply_outer_border(ws, 18,20, 1, 6)

            # INDEX DATA DEFINITION header
            # ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            # ws.cell(row=current_row, column=1, value=f"INDEX DATA DEFINITION").font = Font(bold=True)
            # current_row += 1
            rows_to_merge = [4, 5, 10, 11, 13, 14, 15, 16]
            merge_c_to_f(ws, rows_to_merge, 3, 6)
            ws.cell(row=current_row, column=1, value="PRODUCT DEFINITION").font = Font(bold=True)
            cell = ws.cell(row=5, column=3, value=risk_location)
            cell.alignment = Alignment(wrap_text=True)
            # --- MERGE AND CENTER CELLS AS PER SAMPLE ---
            # Row 1: Merge A1:F1 and center
            for i in [1,2,3,7,8,19,current_row]:
                ws.merge_cells('A'+str(i)+':F'+str(i))
                ws['A'+str(i)].alignment = Alignment(horizontal='center', vertical='center')
                ws['A'+str(i)].font = Font(bold=True)
            current_row += 2

            for i in [4,5,10,11,13,14,15,16]:
                ws.merge_cells('A'+str(i)+':B'+str(i))
                ws['A'+str(i)].alignment = Alignment(horizontal='center', vertical='center')
                ws['A'+str(i)].font = Font(bold=True)

            for i in [13,14,15,16]:
                ws['C'+str(i)].alignment = Alignment(horizontal='left', vertical='center')
            backup_dict = {}
            row_height_dict = {}
            # --- COVER SECTIONS ---
            for i, ds in enumerate(sources):
                cover = covers[i]
                backup_dict["Cover " + str(i+1)] = masters_dict[ds].get('Backup Data', '')            
                cover_info = covers_dict[cover]
                x_val_list = cover_info.get('Variable', '').split('&') if pd.notna(cover_info.get('Variable', '')) else []
                x_def_list = cover_info.get('Variable_for_termsheet', '').split('&') if pd.notna(cover_info.get('Variable_for_termsheet', '')) else []
                x_unit_list = cover_info.get('Variable_unit', '').split('&') if pd.notna(cover_info.get('Variable_unit', '')) else []
                # if "DAY" in cover_info['Unit'].upper():
                #     n = thresholds[i]
                # else:
                #     n = n_days[i]
                date_str = phases[i]
                ordinal = lambda d: f"{d}{'th' if 11<=d%100<=13 else {1:'st',2:'nd',3:'rd'}.get(d%10, 'th')}"
                start, end = [datetime.strptime(d, "%d-%m-%Y") for d in date_str.split('&')]
                formatted_date = f"{ordinal(start.day)} {start.strftime('%B %Y')} to {ordinal(end.day)} {end.strftime('%B %Y')}"
                
                value_list = parse_list(row['x_val'])[i].split('&')
                
                observed_index_raw = str(cover_info.get('Observed_index', ''))
                # Replace each 'N' with the corresponding value
                observed_index = re.sub(r'\bN\b', lambda _: value_list.pop(0), observed_index_raw)
                border_1 = current_row
                source_text = f"{cover_info.get('Cover_name', '')}\nObserved Index: {observed_index}\n{cover_info.get('Claims_payable', '')}"
                ws.cell(row=current_row, column=1, value=f"Cover {i+1}: {source_text}")
                ws.cell(row=current_row, column=4, value=f"Phase Period: {formatted_date}")
                ws.merge_cells('D'+str(current_row)+':F'+str(current_row))
                ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['D'+str(current_row)].font = Font(bold=True)
                apply_outer_border(ws, border_1,border_1, 3, 6)
                current_row += 1
                counter = 0
                for val, definition, unit in zip(x_val_list, x_def_list, x_unit_list):
                    if val != 'NA':
                        to_pick = parse_list(row[val])[i]
                        ws.cell(row=current_row, column=4, value=definition)
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        ws.cell(row=current_row, column=6, value=f"{to_pick} {unit}")
                        current_row += 1
                        counter += 1
                #  a) CYCLONE
                if "CYCLONE" in cover.upper():
                    ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+len(parse_cyclone_entries(cycl_payout[i]))+1))
                    ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    apply_outer_border(ws, border_1,current_row+len(parse_cyclone_entries(cycl_payout[i]))+1, 1, 3)

                    ws.cell(row=current_row, column=4, value="Speed").font = bold_font
                    ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=5, value="Distance").font = bold_font
                    ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=6, value="Payout").font = bold_font
                    ws['F'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    apply_outer_border(ws, current_row,current_row, 3, 6)
                    current_row += 1
                    speed_2 = 0
                    for speed, dist, wt in parse_cyclone_entries(cycl_payout[i]):
                        ws.cell(row=current_row, column=4, value=f">= {smart_convert(speed)} km/hr")
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=5, value=f"<= {smart_convert(dist)} km")
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=6, value=wt * unit_si)
                        if speed_2 != speed:
                            for cyc_col in range(4,7):    
                                ws.cell(row=current_row, column=cyc_col).border = Border(top=Side(style='thin'))    
                        speed_2 = speed
                        current_row += 1
                    apply_outer_border(ws, border_1,current_row, 1, 6)
                    # Maximum Payout
                    ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                    ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                    ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['D'+str(current_row)].font = Font(bold=True)

                    ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                    ws['F'+str(current_row)].font = Font(bold=True)
                    apply_outer_border(ws, current_row,current_row, 3, 6)
                    apply_outer_border(ws, border_1,current_row, 1, 3)
                    row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row+1))
                    current_row += 1

                #  b) VAR cover
                elif "VAR" in cover.upper():
                    if var_types[i] == 'Continuous':
                        ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+4))
                        ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
                        # Base Payout
                        ws.cell(row=current_row, column=4, value="Base Payout")
                        ws.cell(row=current_row, column=6, value=float(var_bpays[i]))
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1
                        # Trigger Level
                        ws.cell(row=current_row, column=4, value="Entry Trigger")
                        ws.cell(row=current_row, column=6, value=f"{smart_convert(var_deds[i])} {cover_info.get('Unit', '')}")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1
                        # Exit Level
                        ws.cell(row=current_row, column=4, value="Exit Trigger")
                        ws.cell(row=current_row, column=6, value=f"{smart_convert(var_deds[i])+smart_convert(var_lvls[i])} {cover_info.get('Unit', '')}")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1
                        # Notional Level
                        ws.cell(row=current_row, column=4, value= f"Notional Payout (per {cover_info.get('Unit', '')})")
                        ws.cell(row=current_row, column=6, value=round((unit_si*float(si_weights[i])- float(var_bpays[i]))/float(var_lvls[i]),0))
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1

                        # Maximum Payout
                        ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)

                        ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                        ws['F'+str(current_row)].font = Font(bold=True)
                        current_row += 1

                        ws.cell(row=current_row, column=1, value="Payout Formula")
                        ws['A'+str(current_row)].alignment = Alignment(wrap_text=True, vertical='center')
                        ws['A'+str(current_row)].font = Font(bold=True)

                        ws.cell(row=current_row, column=2, value="Payout = Min(Base Payout + Max(Observed Index - Entry Trigger,0) x (Notional Payout), Cover Maximum Payout)")
                        ws.merge_cells('B'+str(current_row)+':F'+str(current_row))
                        ws['B'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['B'+str(current_row)].font = Font(italic=True)
                        apply_outer_border(ws, border_1,current_row, 1, 6)
                        apply_outer_border(ws, border_1+1,current_row-2, 3, 6)
                        apply_outer_border(ws, current_row-1,current_row-1, 3, 6)
                        apply_outer_border(ws, border_1,current_row-1, 1, 3)
                        row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row))
                        current_row += 1

                    else:
                        ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+4))
                        ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

                        payout = round(unit_si*float(si_weights[i])*float(var_szs[i])/float(var_lvls[i]),0)

                        ws.cell(row=current_row, column=5, value="Trigger Level").font = bold_font
                        ws.cell(row=current_row, column=6, value="Payout").font = bold_font
                        current_row += 1

                        ws.cell(row=current_row, column=4, value="Entry Trigger").font = bold_font
                        ws.cell(row=current_row, column=5, value=f"{smart_convert(var_deds[i])+smart_convert(var_szs[i])} {cover_info.get('Unit', '')}")
                        ws.cell(row=current_row, column=6, value=payout+float(var_bpays[i]))
                        current_row += 1

                        if "LESS THAN" in cover_info["Claims_payable"].upper():
                            ws.cell(row=current_row, column=4, value="Decrement").font = bold_font
                        else:
                            ws.cell(row=current_row, column=4, value="Increment").font = bold_font
                        ws.cell(row=current_row, column=5, value=f"{smart_convert(var_szs[i])} {cover_info.get('Unit', '')}")
                        ws.cell(row=current_row, column=6, value=payout)
                        current_row += 1

                        ws.cell(row=current_row, column=4, value="Exit Trigger").font = bold_font
                        ws.cell(row=current_row, column=5, value=f"{round(smart_convert(var_lvls[i])+smart_convert(var_deds[i]),0)} {cover_info.get('Unit', '')}")
                        ws.cell(row=current_row, column=6, value=payout*float(var_lvls[i])+float(var_bpays[i]))
                        current_row += 1

                        ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)

                        ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                        ws['F'+str(current_row)].font = Font(bold=True)
                        apply_outer_border(ws, current_row,current_row, 3, 6)
                        apply_outer_border(ws, border_1+1,current_row-1, 3, 6)
                        apply_outer_border(ws, border_1,current_row, 1, 3)
                        row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row+1))
                        current_row += 1
                    
                #  c) simple strike‐based
                else:
                    if strikes and strikes[i] != 'NA':
                        levels = strikes[i].split('&')
                        wts   = strike_wts[i].split('&')
                        # header
                        ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+1+len(levels)))
                        ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                        ws.cell(row=current_row, column=5, value="Trigger Level").font = bold_font
                        ws.cell(row=current_row, column=6, value="Payout").font = bold_font
                        current_row += 1

                        for j, (lvl, wt) in enumerate(zip(levels, wts), start=1):
                            if j == len(levels):
                                ws.cell(row=current_row, column=4, value=f"Strike {j} (Exit)").font = bold_font
                            else:
                                ws.cell(row=current_row, column=4, value=f"Strike {j}").font = bold_font
                            ws.cell(row=current_row, column=5, value=f"{lvl} {cover_info.get('Unit', '')}")
                            ws.cell(row=current_row, column=6, value=float(wt) * unit_si*float(si_weights[i]))
                            current_row += 1
                        
                        ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)

                        ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                        ws['F'+str(current_row)].font = Font(bold=True)
                        apply_outer_border(ws, border_1,current_row, 1, 6)
                        apply_outer_border(ws, current_row,current_row, 3, 6)
                        apply_outer_border(ws, border_1+1,current_row-1, 3, 6)
                        apply_outer_border(ws, border_1,current_row, 1, 3)
                        row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row+1))
                        current_row += 1
                    
                ds = ds.strip()
                if ds not in masters_dict:
                    continue

                master_info = masters_dict[ds]
                
                # # INDEX DATA NAME
                # ws.cell(row=current_row, column=1, value=f"COVER {i+1} NAME").font = Font(bold=True)
                # name_value = master_info.get('Name', '')
                # trimmed_name = name_value.split(':')[0] if ':' in name_value else name_value
                # merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                # ws.cell(row=current_row, column=3, value=trimmed_name).alignment = Alignment(wrap_text=True)

                # current_row += 1
                
                # INDEX DATA SOURCE
                ws.cell(row=current_row, column=1, value=f"Cover {i+1} Data Source").font = Font(bold=True)
                ws.merge_cells('A'+str(current_row)+':B'+str(current_row))
                ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                source_text = f"{master_info.get('Source', '')}\nWeblink to Access Data: {master_info.get('Website', '')}\n{master_info.get('Name', '')}"
                merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                ws.cell(row=current_row, column=3, value=source_text).alignment = Alignment(wrap_text=True)
                apply_outer_border(ws, current_row,current_row+1, 1, 6)
                current_row += 1
                
                # # INDEX DATA BACKUP
                # ws.cell(row=current_row, column=1, value=f"COVER {i+1} DATA BACKUP").font = Font(bold=True)
                # merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                # ws.cell(row=current_row, column=3, value=master_info.get('Backup Data', '')).alignment = Alignment(wrap_text=True)
                # current_row += 1
                
                # INDEX DATA GEO-REFERENCE
                ws.cell(row=current_row, column=1, value=f"Cover {i+1} Data Geo-Reference").font = Font(bold=True)
                ws.merge_cells('A'+str(current_row)+':B'+str(current_row))
                ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                try:
                    lat, long = ref_lat_long[i].split('&')
                except Exception:
                    lat = "NA"
                    long = "NA"
                formatted_coord = f"{lat}°N {long}°E"            
                ws.cell(row=current_row, column=3, value=formatted_coord).alignment = Alignment(wrap_text=True)
                current_row += 2  # Add spacing between blocks

            
            ws.cell(row=current_row, column=1, value=f"Total claim payout under the policy limited to ₹{smart_convert(unit_si)} per {unit_type.lower()}")
            ws.merge_cells('A'+str(current_row)+':F'+str(current_row))
            ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
            ws['A'+str(current_row)].font = Font(bold=True, italic=True)
            apply_outer_border(ws, current_row,current_row, 1, 6)            
            current_row += 2
            ws.cell(row=current_row, column=1, value="This policy covers only losses attributed to the perils covered during the specified period. Losses due to any other perils are not covered")
            ws.merge_cells('A'+str(current_row)+':F'+str(current_row))
            ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
            ws['A'+str(current_row)].font = Font(italic=True)
            apply_outer_border(ws, current_row,current_row, 1, 6)
            current_row += 3
            grouped = defaultdict(list)
            for cover, message in backup_dict.items():
                grouped[message].append(cover)

            # Generate the summary string
            summary_parts = []
            for message, covers in grouped.items():
                cover_list = ', '.join(covers)
                summary_parts.append(f"For {cover_list}\n{message}")
            final_summary = '\n'.join(summary_parts)            
            ws.cell(row=current_row, column=1, value="Backup Data Source")
            ws.merge_cells('A'+str(current_row)+':B'+str(current_row))
            ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
            ws['A'+str(current_row)].font = Font(bold=True)

            ws.cell(row=current_row, column=3, value=final_summary)
            ws.merge_cells('C'+str(current_row)+':F'+str(current_row))
            ws['C'+str(current_row)].alignment = Alignment(wrap_text=True)
            apply_outer_border(ws, current_row,current_row, 1, 6)
            current_row += 1

            # --- BOLD HEADINGS ---
            bold_rows = [4, 5, 10, 11, 13, 14, 15, 16]
            for r in bold_rows:
                cell = ws.cell(row=r, column=1)
                if cell.value not in [None, '']:
                    cell.font = Font(bold=True)

            widths = [10.14, 25.14, 20.14, 21.14, 40.86, 21.14]  # Custom widths for columns A–F
            for i, width in enumerate(widths, start=1):
                col_letter = get_column_letter(i)  # Convert column number to letter (1 → 'A', etc.)
                ws.column_dimensions[col_letter].width = width

            rupee_format = u'₹#,##,##0.00'

            for row in ws.iter_rows():
                for cell in row:
                    old_font = cell.font or Font()
                    cell.font = Font(
                        name="Times New Roman",
                        size=12,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color=old_font.color
                    )
                    try:
                        # Try to interpret the cell value as a number
                        float(cell.value)
                        cell.number_format = rupee_format
                    except (TypeError, ValueError):
                        # Skip if it's not a number or cannot be converted to float
                        pass        
            ws.sheet_view.zoomScale = 85
            chars_per_line = 85  # Average number of characters per line
            default_row_height = 17  # Default row height in pixels

            # Set all row heights to 17px initially
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = default_row_height

            # Loop through all rows and columns in the worksheet
            for row_num in range(7, ws.max_row + 1):  # Loop through all rows
                max_chars_in_row = 0
                # Loop through all columns in the row
                for col_num in range(1, ws.max_column + 1):  # Loop through all columns
                    cell = ws.cell(row=row_num, column=col_num)
                    if col_num in (4,5,6):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell_data = str(cell.value) if cell.value else ""
                    
                    # Update the max_chars_in_row for the row based on the current cell
                    max_chars_in_row = max(max_chars_in_row, len(cell_data))

                # Estimate the number of lines needed for the row
                num_lines = math.ceil(max_chars_in_row / chars_per_line)

                # Calculate the row height based on the number of lines
                row_height = default_row_height * num_lines

                # Set the row height for this row
                ws.row_dimensions[row_num].height = row_height
            set_dynamic_row_heights(ws, row_height_dict)
            # Set the print area (columns A to F, full used row range)
            last_row = ws.max_row
            ws.print_area = f"A1:F{last_row}"
            # Page layout: Fit columns on one page, allow rows to go to multiple pages
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0  # 0 means "automatic" (multiple pages tall)
            ws.page_setup.fitToWidth = 1             
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[5].height = 17 * line_count
            ws.print_title_rows = '1:7'
            # for row in ws.iter_rows(min_row=1, max_row=6):
            #     for cell in row:
            #         if cell.value == "RISK LOCATION":
            #             target_cell = ws.cell(row=cell.row, column=3)
            #             target_cell.alignment = Alignment(wrap_text=True, vertical='top')
            wb.save(output_file)
            save_excel_sheet_as_pdf(output_file, 'Draft Termsheet', output_pdf, './working_folder/Termsheet/logo.png', formatted_time)
            if os.path.exists(output_file):
                os.remove(output_file)
    else:
        cols_to_check = ['Num_Risk_Units', 'Covers', 'Cyclone_Speed_Dist_PayoutWt']
        df_uploader['strike_pattern'] = df_uploader['Strikes'].apply(parse_strike_counts)
        df_uploader['combo_key'] = df_uploader[cols_to_check + ['strike_pattern']].astype(str).agg('|'.join, axis=1)

        unique_combos = df_uploader['combo_key'].unique()

        # Report differences
        differences_found = False
        for col in cols_to_check:
            if df_uploader[col].nunique() > 1:
                differences_found = True
                # print(f"\nMultiple values in column '{col}':")
                # print(df_uploader[col].drop_duplicates().tolist())

        if len(df_uploader['strike_pattern'].unique()) > 1:
            differences_found = True
            # print("\nMultiple strike structures found:")
            # print(df_uploader['strike_pattern'].drop_duplicates().tolist())

        # Prompt user for cohort names if differences exist
        if differences_found:
            print("\nPlease assign cohort names for each unique combination:")
            combo_to_cohort = {}
            for i, combo in enumerate(unique_combos, 1):
                print(f"\nCombination {i}:")
                key_parts = combo.split('|')
                for col_name, val in zip(cols_to_check + ['strike_pattern'], key_parts):
                    print(f"{col_name}: {val}")
                cohort_name = input(f"Input Cohort name for Combination {i}: ")
                combo_to_cohort[combo] = cohort_name

            # Assign the cohort names to each row
            df_uploader['Cohort'] = df_uploader['combo_key'].map(combo_to_cohort)
            print("\nCohort names assigned to DataFrame.")
        else:
            df_uploader['Cohort'] = "NA"
        cohort_dfs = {cohort: group_df.reset_index(drop=True) for cohort, group_df in df_uploader.groupby('Cohort')}
        for cohort_name, df_uploader in cohort_dfs.items():
            now = datetime.now()
            base_time_str = now.strftime("%Y-%m-%d_%H-%M-%S")
            base_microsecond = now.microsecond
            formatted_time = base_time_str + "-" + str(base_microsecond)
            annex = "Refer Annexure"
            output_file = f'Draft Termsheet {str(formatted_time)}.xlsx'
            # Read data
            df = df_uploader.copy()
            # Convert each to dictionary
            masters_dict = first_table_df.set_index(first_table_df.columns[0]).to_dict(orient='index')
            covers_dict = second_table_df.set_index(second_table_df.columns[0]).to_dict(orient='index')
            row = df_uploader.iloc[0]
            unit_type = row['Unit_Type'].upper()
            if df['loc_name'].nunique() == 1:
                risk_location, line_count = format_risk_location(row['loc_name'])
            else:
                risk_location = annex
                line_count = 1
            risk_start = format_date(pd.to_datetime(row['RSD'], format="%d-%m-%Y"))
            risk_end = format_date(pd.to_datetime(row['RED'], format="%d-%m-%Y"))
            net_prem = float(row['Unit_Net_Prem'])
            gst = net_prem * float(row['GST(%)'])
            gross = net_prem + gst
            n_units = int(row['Num_Risk_Units'])
            unit_si = float(row['Unit_SI'])
            
            cover_list = parse_list(row['Covers'])
            cover_mapping = {f"Cover {i+1}": cover for i, cover in enumerate(cover_list)}
            cover_list = list(cover_mapping.keys())
            # if "CYCLONE" in cover_list:
            #     cyclone = 1
            # else:
            #     cyclone = 0
            # cover_list.remove("CYCLONE")
            df.drop(columns=['strike_pattern', 'combo_key'],inplace=True)
            param_cols = [col for col in df.columns if col not in [
                'termsheet_id','Risk_Details', 'loc_name', 'RSD', 'RED', 'Unit_Type', 'Unit_SI',
                'Unit_Net_Prem', 'GST(%)', 'Num_Risk_Units', 'Covers', 'Cyclone_Speed_Dist_PayoutWt'
            ]]
            for cover_idx, risk in enumerate(cover_list):
                if risk != "CYCLONE":
                    for param in param_cols:
                        # Split the parameter string into a list
                        param_values = df[param].str.strip('[]').str.split(';')
                        # Create a new column for this cover and parameter
                        df[f"{risk}_{param}"] = param_values.apply(lambda x: x[cover_idx] if len(x) > cover_idx else None)
            
            unique_values = set()
            for val in second_table_df['Variable'].iloc[1:].dropna():
                items = str(val).split('&')
                unique_values.update([item for item in items if item != 'NA'])

            result = list(unique_values)
            
            drop_list = ['Data_Sources','SI_Weights', 'VAR_deductible', 'VAR_levels', 
                        'VAR_level_size', 'VAR_basepay_wt', 'VAR_payout_type','Phase_Dates','Ref_Lat_Lon'] 
            drop_list = drop_list + result
            df.drop(drop_list,axis=1,inplace=True, errors='ignore')

            # Define columns to split and their suffixes
            fields = ['Strikes', 'Strike_SI_Weights']

            for field in fields:
                for risk in cover_list:
                    if risk != "CYCLONE":
                        col = f"{risk}_{field}"

                        if col in df.columns and df[col].notna().any() and df[col].astype(str).str.strip().ne('').any():
                            splits = df[col].astype(str).str.split('&', expand=True)
                            for i in range(splits.shape[1]):
                                df[f"{col}_{i+1}"] = splits[i]
                        df.drop(f"{risk}_{field}",axis=1,inplace=True)
                        
            # parse all the list‐columns
            ref_lat_long = parse_list(row['Ref_Lat_Lon'])
            covers      = parse_list(row['Covers'])
            sources     = parse_list(row['Data_Sources'])
            phases      = parse_list(row['Phase_Dates'])
            si_weights  = parse_list(row['SI_Weights'])
            strikes     = parse_list(row['Strikes'])
            strike_wts  = parse_list(row['Strike_SI_Weights'])
            var_deds    = parse_list(row['VAR_deductible'])
            var_lvls    = parse_list(row['VAR_levels'])
            var_szs     = parse_list(row['VAR_level_size'])
            var_bpays   = parse_list(row['VAR_basepay_wt'])
            var_types   = parse_list(row['VAR_payout_type'])
            cycl_payout = parse_list(row['Cyclone_Speed_Dist_PayoutWt'])
            # Build data as per the sample
            draft_data = [
                ['<<Name of Product>> (UIN: )', None, None, None, None, None],           # 1
                [None, None, None, None, None, None],                                    # 2
                ['RISK DETAILS', None, row['Risk_Details'], None, None, None],           # 3
                ['RISK LOCATION', None, None, None, None, None],                # 4
                [None, None, None, None, None, None],                                    # 5
                [None, None, None, None, None, None],                                    # 6
                [f'PREMIUM & COVERAGE SUMMARY (PER {unit_type})', None, None, None, None, None], # 7
                [None, None, None, None, None, None],                                    # 8
                ['RISK START DATE', None, risk_start, None, None, None],                 # 9
                ['RISK END DATE', None, risk_end, None, None, None],                     # 10
                [None, None, None, None, None, None],                                    # 11
                ['SUM INSURED', None, row['Unit_SI'], None, None, None],                 # 12
                ['PREMIUM AMOUNT', None, round(net_prem,4), None, None, None],                    # 13
                ['GST', None, round(gst,4), None, None, None],                              # 14
                ['GROSS PREMIUM', None, round(gross,4), None, None, None],                        # 15
                [None, None, None, None, None, None],                                    # 16
            ]
            columns = [
                '<<Name of Insurance Company>>', 'Unnamed: 1', 'Unnamed: 2',
                'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 5'
            ]
            df_draft = pd.DataFrame(draft_data, columns=columns)
            df_draft.to_excel(output_file, sheet_name='Draft Termsheet', index=False)
            # Now apply formatting with openpyxl
            wb = load_workbook(output_file)
            if 'Annexure' not in wb.sheetnames:
                wb.create_sheet('Annexure')
                wb.save(output_file)

            ws = wb['Draft Termsheet']
            clear_all_borders(ws)
            current_row = 19

            apply_outer_border(ws, 1,2, 1, 6)
            apply_outer_border(ws, 3,6, 1, 6)
            apply_outer_border(ws, 7,17, 1, 6)
            apply_outer_border(ws, 18,20, 1, 6)
            # INDEX DATA DEFINITION header
            # ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            # ws.cell(row=current_row, column=1, value=f"INDEX DATA DEFINITION").font = Font(bold=True)
            # current_row += 1
            rows_to_merge = [4, 5, 10, 11, 13, 14, 15, 16]
            merge_c_to_f(ws, rows_to_merge, 3, 6)
            # for idx, ds in enumerate(sources,1):
            ws.cell(row=current_row, column=1, value="PRODUCT DEFINITION").font = Font(bold=True)
            cell = ws.cell(row=5, column=3, value=risk_location)
            cell.alignment = Alignment(wrap_text=True)
            # --- MERGE AND CENTER CELLS AS PER SAMPLE ---
            # Row 1: Merge A1:F1 and center
            for i in [1,2,7,8,19,current_row]:
                ws.merge_cells('A'+str(i)+':F'+str(i))
                ws['A'+str(i)].alignment = Alignment(horizontal='center', vertical='center')
                ws['A'+str(i)].font = Font(bold=True)
            current_row += 2

            for i in [4,5,10,11,13,14,15,16]:
                ws.merge_cells('A'+str(i)+':B'+str(i))
                ws['A'+str(i)].alignment = Alignment(horizontal='center', vertical='center')
                ws['A'+str(i)].font = Font(bold=True)

            for i in [13,14,15,16]:
                ws['C'+str(i)].alignment = Alignment(horizontal='left', vertical='center')

            annex_2 = "N(Refer Annexure)"
            backup_dict = {}
            row_height_dict = {}
            # --- COVER SECTIONS ---
            for i, ds in enumerate(sources):
                cover = covers[i]
                backup_dict["Cover " + str(i+1)] = masters_dict[ds].get('Backup Data', '')            
                cover_info = covers_dict[cover]
                cover_def = cover
                x_val_list = cover_info.get('Variable', '').split('&') if pd.notna(cover_info.get('Variable', '')) else []
                x_def_list = cover_info.get('Variable_for_termsheet', '').split('&') if pd.notna(cover_info.get('Variable_for_termsheet', '')) else []
                x_unit_list = cover_info.get('Variable_unit', '').split('&') if pd.notna(cover_info.get('Variable_unit', '')) else []
                cover = cover_list[i]
                for col in df.columns:
                    try:
                        df[col] = df[col].astype(float)
                    except (ValueError, TypeError):
                        pass
                # if df[cover+'_N_Days'].nunique() == 1:
                #     n = n_days[i]
                #     df.drop(cover+'_N_Days',axis=1,inplace=True)
                # else:
                #     n = annex_2
                # if "DAY" in cover_info['Unit'].upper():
                #     if df[cover+'_Threshold'].nunique() == 1:
                #         n = thresholds[i]
                #         df.drop(cover+'_Threshold',axis=1,inplace=True)
                #     else:
                #         n = annex_2
                
                date_str = phases[i]
                ordinal = lambda d: f"{d}{'th' if 11<=d%100<=13 else {1:'st',2:'nd',3:'rd'}.get(d%10, 'th')}"
                start, end = [datetime.strptime(d, "%d-%m-%Y") for d in date_str.split('&')]
                formatted_date = f"{ordinal(start.day)} {start.strftime('%B %Y')} to {ordinal(end.day)} {end.strftime('%B %Y')}"

                # x_val = re.sub(r'\bN\b', str(n),str(cover_info.get('Variable', '')))
                observed_index = str(cover_info.get('Observed_index_annex', ''))
                # Replace standalone 'N' only (e.g., 'N', not 'IN' or 'NEXT')
                
                # observed_index = re.sub(r'\bN\b', str(n), observed_index)
                border_1 = current_row
                source_text = f"{cover_info.get('Cover_name', '')}\nObserved Index: {observed_index}\n{cover_info.get('Claims_payable', '')}"
                ws.cell(row=current_row, column=1, value=f"Cover {i+1}: {source_text}")
                # ws.write_rich_string(current_row, 0, b, f"Cover {i+1}: ", b, f"{cover_info.get('Cover_name', '')}\n", i_style, f"Observed Index: {observed_index}\n{cover_info.get('Claims_payable', '')}")
                if df[cover+'_Phase_Dates'].nunique() == 1:
                    ws.cell(row=current_row, column=4, value=f"Phase Period: {formatted_date}")
                    df.drop(cover+'_Phase_Dates',axis=1,inplace=True)
                else:
                    ws.cell(row=current_row, column=4, value=f"Phase Period: {annex}")
                ws.merge_cells('D'+str(current_row)+':F'+str(current_row))
                ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['D'+str(current_row)].font = Font(bold=True)
                apply_outer_border(ws, border_1,border_1, 3, 6)
                current_row += 1
                counter = 0
                for val, definition, unit in zip(x_val_list, x_def_list, x_unit_list):
                    to_pick = parse_list(row[val])[i]
                    if val != 'NA':
                        if df[cover+'_'+val].nunique() == 1:
                            pass
                        else:
                            to_pick = annex
                        ws.cell(row=current_row, column=4, value=definition)
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        ws.cell(row=current_row, column=6, value=f"{to_pick} {unit}")
                        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1
                        counter += 1

                #  a) CYCLONE
                if "CYCLONE" in cover_def.upper():
                    ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+len(parse_cyclone_entries(cycl_payout[i]))+1))
                    # ws['A'+str(current_row-1-counter)].alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    apply_outer_border(ws, border_1,current_row+len(parse_cyclone_entries(cycl_payout[i]))+1, 1, 3)

                    ws.cell(row=current_row, column=4, value="Speed").font = bold_font
                    ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=5, value="Distance").font = bold_font
                    ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=current_row, column=6, value="Payout").font = bold_font
                    ws['F'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    apply_outer_border(ws, current_row,current_row, 3, 6)
                    current_row += 1
                    speed_2 = 0
                    for speed, dist, wt in parse_cyclone_entries(cycl_payout[i]):
                        ws.cell(row=current_row, column=4, value=f">= {smart_convert(speed)} km/hr")
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=5, value=f"<= {smart_convert(dist)} km")
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=6, value=wt * unit_si)
                        if speed_2 != speed:
                            for cyc_col in range(4,7):    
                                ws.cell(row=current_row, column=cyc_col).border = Border(top=Side(style='thin'))    
                        speed_2 = speed
                        current_row += 1
                    # apply_outer_border(ws, border_1+2,current_row-1, 3, 6)
                    apply_outer_border(ws, border_1,current_row, 1, 6)

                    # Maximum Payout
                    ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                    ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                    ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                    ws['D'+str(current_row)].font = Font(bold=True)

                    ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                    ws['F'+str(current_row)].font = Font(bold=True)
                    apply_outer_border(ws, current_row,current_row, 3, 6)
                    apply_outer_border(ws, border_1,current_row, 1, 3)
                    row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row+1))
                    current_row += 1
                #  b) VAR cover
                elif "VAR" in cover_def.upper():
                    if var_types[i] == 'Continuous':
                        ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+4))
                        ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
                        # Base Payout
                        ws.cell(row=current_row, column=4, value="Base Payout")
                        if df[cover+'_VAR_basepay_wt'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=float(var_bpays[i]))
                        else:
                            ws.cell(row=current_row, column=6, value=annex)
        
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1

                        # Trigger Level
                        ws.cell(row=current_row, column=4, value="Entry Trigger")
                        if df[cover+'_VAR_deductible'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=f"{smart_convert(var_deds[i])} {cover_info.get('Unit', '')}")
                        else:
                            ws.cell(row=current_row, column=6, value=annex)
                            
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1
                        # Exit Level

                        ws.cell(row=current_row, column=4, value="Exit Trigger")
                        df[cover+' Exit Trigger'] = df[cover+'_VAR_deductible']+ df[cover+'_VAR_levels']
                        if df[cover+' Exit Trigger'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=f"{smart_convert(var_deds[i])+smart_convert(var_lvls[i])} {cover_info.get('Unit', '')}")
                        else:
                            ws.cell(row=current_row, column=6, value=annex)


                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1
                        # Notional Level
                        ws.cell(row=current_row, column=4, value= f"Notional Payout (per {cover_info.get('Unit', '')})")

                        df[cover+' Notional Payout'] = (df['Unit_SI']*df[cover+'_SI_Weights']-float(var_bpays[i]))/df[cover+'_VAR_levels']
                        if df[cover+' Notional Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=round((unit_si*float(si_weights[i])- float(var_bpays[i]))/float(var_lvls[i]),0))
                        else:
                            ws.cell(row=current_row, column=6, value=annex)
                        

                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)
                        current_row += 1

                        # Maximum Payout
                        ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)

                        df[cover+' Maximum Payout'] = df['Unit_SI']*df[cover+'_SI_Weights']
                        if df[cover+' Maximum Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                        else:
                            ws.cell(row=current_row, column=6, value=annex)

                        ws['F'+str(current_row)].font = Font(bold=True)
                        current_row += 1

                        ws.cell(row=current_row, column=1, value="Payout Formula")
                        ws['A'+str(current_row)].alignment = Alignment(wrap_text=True, vertical='center')
                        ws['A'+str(current_row)].font = Font(bold=True)

                        ws.cell(row=current_row, column=2, value="Payout = Min(Base Payout + Max(Observed Index - Entry Trigger,0) x (Notional Payout), Cover Maximum Payout)")
                        ws.merge_cells('B'+str(current_row)+':F'+str(current_row))
                        ws['B'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['B'+str(current_row)].font = Font(italic=True)
                        apply_outer_border(ws, border_1,current_row, 1, 6)
                        apply_outer_border(ws, border_1+1,current_row-2, 3, 6)
                        apply_outer_border(ws, current_row-1,current_row-1, 3, 6)
                        apply_outer_border(ws, border_1,current_row-1, 1, 3)
                        row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row))
                        current_row += 1

                    else:
                        ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+4))
                        ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

                        payout = round(unit_si*float(si_weights[i])*float(var_szs[i])/float(var_lvls[i]),0)

                        ws.cell(row=current_row, column=5, value="Trigger Level").font = bold_font
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=6, value="Payout").font = bold_font
                        ws['F'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1


                        ws.cell(row=current_row, column=4, value="Entry Trigger").font = bold_font
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')

                        df[cover+' Entry Trigger'] = df[cover+'_VAR_deductible'] + df[cover+'_VAR_level_size']
                        if df[cover+' Entry Trigger'].nunique() == 1:
                            ws.cell(row=current_row, column=5, value=f"{smart_convert(var_deds[i])+1} {cover_info.get('Unit', '')}")
                        else:
                            ws.cell(row=current_row, column=5, value=annex)
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')

                        df[cover+' Entry Trigger Payout'] = (df['Unit_SI']*df[cover+'_SI_Weights']*df[cover+'_VAR_level_size'])/df[cover+'_VAR_levels'] + df[cover+'_VAR_basepay_wt']
                        if df[cover+' Entry Trigger Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=payout)
                        else:
                            ws.cell(row=current_row, column=6, value=annex)

                        current_row += 1
                        if "LESS THAN" in cover_info["Claims_payable"].upper():
                            ws.cell(row=current_row, column=4, value="Decrement").font = bold_font
                        else:
                            ws.cell(row=current_row, column=4, value="Increment").font = bold_font
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')

                        if df[cover+'_VAR_level_size'].nunique() == 1:
                            ws.cell(row=current_row, column=5, value=f"{smart_convert(var_szs[i])} {cover_info.get('Unit', '')}")
                        else:
                            ws.cell(row=current_row, column=5, value=annex)
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        df[cover+' Increment Payout'] = (df['Unit_SI']*df[cover+'_SI_Weights']*df[cover+'_VAR_level_size'])/df[cover+'_VAR_levels']
                        if df[cover+' Increment Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=payout)
                        else:
                            ws.cell(row=current_row, column=6, value=annex)

                        current_row += 1

                        ws.cell(row=current_row, column=4, value="Exit Trigger").font = bold_font
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        df[cover+' Exit Trigger'] = df[cover+'_VAR_levels']+df[cover+'_VAR_deductible']
                        if df[cover+' Exit Trigger'].nunique() == 1:
                            ws.cell(row=current_row, column=5, value=f"{round(smart_convert(var_lvls[i])+smart_convert(var_deds[i]),0)} {cover_info.get('Unit', '')}")
                        else:
                            ws.cell(row=current_row, column=5, value=annex)
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')

                        df[cover+' Exit Trigger Payout'] = df[cover+'_VAR_levels']*df[cover+' Increment Payout'] + df[cover+'_VAR_basepay_wt']
                        if df[cover+' Exit Trigger Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=payout*float(var_lvls[i]))
                        else:
                            ws.cell(row=current_row, column=6, value=annex)
                        current_row += 1

                        ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)

                        df[cover+' Maximum Payout'] = df['Unit_SI']*df[cover+'_SI_Weights']
                        if df[cover+' Maximum Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                        else:
                            ws.cell(row=current_row, column=6, value=annex)
                        ws['F'+str(current_row)].font = Font(bold=True)
                        apply_outer_border(ws, current_row,current_row, 3, 6)
                        apply_outer_border(ws, border_1+1,current_row-1, 3, 6)
                        apply_outer_border(ws, border_1,current_row, 1, 3)
                        row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row+1))
                        current_row += 1
                    
                #  c) simple strike‐based
                else:
                    if strikes and strikes[i] != 'NA':
                        levels = strikes[i].split('&')
                        wts   = strike_wts[i].split('&')
                        # header
                        ws.merge_cells('A'+str(current_row-1-counter)+':C'+str(current_row+1+len(levels)))
                        ws.cell(row=current_row-1-counter, column=1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                        ws.cell(row=current_row, column=5, value="Trigger Level").font = bold_font
                        ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=current_row, column=6, value="Payout").font = bold_font
                        ws['F'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1

                        for j, (lvl, wt) in enumerate(zip(levels, wts), start=1):
                            if j == len(levels):
                                ws.cell(row=current_row, column=4, value=f"Strike {j} (Exit)").font = bold_font
                            else:
                                ws.cell(row=current_row, column=4, value=f"Strike {j}").font = bold_font
                            ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                            df[cover+f' Strike {j} Payout'] = df[cover+f'_Strike_SI_Weights_{j}']*unit_si*df[cover+'_SI_Weights']
                            if df[cover+f'_Strikes_{j}'].nunique() == 1 and df[cover+f'_Strike_SI_Weights_{j}'].nunique() == 1:
                                ws.cell(row=current_row, column=5, value=f"{lvl} {cover_info.get('Unit', '')}")
                                ws.cell(row=current_row, column=6, value=float(wt) * unit_si*float(si_weights[i]))
                            else:
                                ws.cell(row=current_row, column=5, value=annex)
                                ws.cell(row=current_row, column=6, value=annex)
                            ws['E'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                            current_row += 1
                            
                        ws.cell(row=current_row, column=4, value="Cover Maximum Payout")
                        ws.merge_cells('D'+str(current_row)+':E'+str(current_row))
                        ws['D'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                        ws['D'+str(current_row)].font = Font(bold=True)

                        df[cover+' Maximum Payout'] = df['Unit_SI']*df[cover+'_SI_Weights']
                        if df[cover+' Maximum Payout'].nunique() == 1:
                            ws.cell(row=current_row, column=6, value=unit_si*float(si_weights[i]))
                        else:
                            ws.cell(row=current_row, column=6, value=annex)

                        ws['F'+str(current_row)].font = Font(bold=True)
                        apply_outer_border(ws, border_1,current_row, 1, 6)
                        apply_outer_border(ws, current_row,current_row, 3, 6)
                        apply_outer_border(ws, border_1+1,current_row-1, 3, 6)
                        apply_outer_border(ws, border_1,current_row, 1, 3)
                        row_height_dict[f"Cover {i+1}"] = list(range(border_1,current_row+1))
                        current_row += 1
                    
                ds = ds.strip()
                if ds not in masters_dict:
                    continue

                master_info = masters_dict[ds]
                
                # # INDEX DATA NAME
                # ws.cell(row=current_row, column=1, value=f"INDEX DATA {i+1} NAME").font = Font(bold=True)
                # name_value = master_info.get('Name', '')
                # trimmed_name = name_value.split(':')[0] if ':' in name_value else name_value
                # merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                # ws.cell(row=current_row, column=3, value=trimmed_name).alignment = Alignment(wrap_text=True)
                
                # INDEX DATA SOURCE
                ws.cell(row=current_row, column=1, value=f"Cover {i+1} Data Source").font = Font(bold=True)
                ws.merge_cells('A'+str(current_row)+':B'+str(current_row))
                ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                source_text = f"{master_info.get('Source', '')}\nWeblink to Access Data: {master_info.get('Website', '')}\n{master_info.get('Name', '')}"
                merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                ws.cell(row=current_row, column=3, value=source_text).alignment = Alignment(wrap_text=True)
                apply_outer_border(ws, current_row,current_row+1, 1, 6)
                current_row += 1
                # # INDEX DATA BACKUP
                # ws.cell(row=current_row, column=1, value=f"INDEX DATA {i+1} BACKUP").font = Font(bold=True)
                # merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                # ws.cell(row=current_row, column=3, value=master_info.get('Backup Data', '')).alignment = Alignment(wrap_text=True)
                # current_row += 1
                
                # INDEX DATA GEO-REFERENCE
                ws.cell(row=current_row, column=1, value=f"Cover {i+1} Data Geo-Reference").font = Font(bold=True)
                ws.merge_cells('A'+str(current_row)+':B'+str(current_row))
                ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
                merge_center(ws,'C'+ str(current_row),'f'+ str(current_row))
                try:
                    lat, long = ref_lat_long[i].split('&')
                except Exception:
                    lat = "NA"
                    long = "NA"
                formatted_coord = f"{lat}°N {long}°E"            
                if df[cover+'_Ref_Lat_Lon'].nunique() == 1:
                    ws.cell(row=current_row, column=3, value=formatted_coord).alignment = Alignment(wrap_text=True)
                else:
                    ws.cell(row=current_row, column=3, value=annex).alignment = Alignment(wrap_text=True)
                current_row += 2 

            
            ws.cell(row=current_row, column=1, value=f"Total claim payout under the policy limited to ₹{smart_convert(unit_si)} per {unit_type.lower()}")
            ws.merge_cells('A'+str(current_row)+':F'+str(current_row))
            ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
            ws['A'+str(current_row)].font = Font(bold=True, italic=True)
            apply_outer_border(ws, current_row,current_row, 1, 6)            
            current_row += 2
            ws.cell(row=current_row, column=1, value="This policy covers only losses attributed to the perils covered during the specified period. Losses due to any other perils are not covered")
            ws.merge_cells('A'+str(current_row)+':F'+str(current_row))
            ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
            ws['A'+str(current_row)].font = Font(italic=True)
            apply_outer_border(ws, current_row,current_row, 1, 6)
            current_row += 3
            grouped = defaultdict(list)
            for cover, message in backup_dict.items():
                grouped[message].append(cover)

            # Generate the summary string
            summary_parts = []
            for message, covers in grouped.items():
                cover_list = ', '.join(covers)
                summary_parts.append(f"For {cover_list}\n{message}")

            final_summary = '\n'.join(summary_parts)            
            ws.cell(row=current_row, column=1, value="Backup Data Source")
            ws.merge_cells('A'+str(current_row)+':B'+str(current_row))
            ws['A'+str(current_row)].alignment = Alignment(horizontal='center', vertical='center')
            ws['A'+str(current_row)].font = Font(bold=True)

            ws.cell(row=current_row, column=3, value=final_summary)
            ws.merge_cells('C'+str(current_row)+':F'+str(current_row))
            ws['C'+str(current_row)].alignment = Alignment(wrap_text=True)
            apply_outer_border(ws, current_row,current_row, 1, 6)
            current_row += 1

            bold_rows = [4, 5, 10, 11, 13, 14, 15, 16]
            for r in bold_rows:
                cell = ws.cell(row=r, column=1)
                if cell.value not in [None, '']:
                    cell.font = Font(bold=True)

            widths = [10.14, 25.14, 20.14, 21.14, 40.86, 21.14]  # Custom widths for columns A–F
            for i, width in enumerate(widths, start=1):
                col_letter = get_column_letter(i)  # Convert column number to letter (1 → 'A', etc.)
                ws.column_dimensions[col_letter].width = width
                wb.save(output_file)

            fields = ['Strikes', 'Strike_SI_Weights']
            df.drop(fields, axis=1,inplace=True)
            df = df.replace(r'^\s*$', np.nan, regex=True)
            df = clean_column_names(df)
            # Step 1: Split Location into components
            location_parts = df['Location'].str.strip('[]').str.split(';', expand=True)

            # Step 2: Format lat and lon into a single string
            df['Geo-Coordinates'] = location_parts[0].str.split('_').apply(lambda x: f"{x[0]}°N {x[1]}°E" if len(x) == 2 and x[0] != "NA" and x[1] != "NA" else "NA")
            # Step 3: Assign remaining fields
            df['Pincode'] = location_parts[1]
            df['Block'] = location_parts[2]
            df['District'] = location_parts[3]
            df['State'] = location_parts[4]
            df['Country'] = location_parts[5]
            df_replaced = df.replace("NA", pd.NA)

            # Drop columns where all values are NA
            df = df_replaced.dropna(axis=1, how='all')

            # (Optional) Replace back if you want to keep "NA" elsewhere
            df = df.fillna("NA")
            
            df = df.loc[:, (df.nunique(dropna=False) > 1) | (df.columns == 'Termsheet Id')]
            df.rename(columns={"Termsheet Id": "Termsheet ID"}, inplace=True)
            if cohort_name != "NA":
                df['Cohort'] = cohort_name
            rupee_format = u'₹#,##,##0.00'
            term_id = [c for c in df.columns if c == 'Termsheet ID']
            loc_cols = [c for c in df.columns if any(k in c for k in ['Geo', 'Pincode', 'Block', 'District', 'State', 'Country'])]
            covers = {}
            for c in df.columns:
                m = re.match(r'^Cover (\d+)', c, re.I)
                if m:
                    covers.setdefault(m.group(1), []).append(c)

            cover_cols = [col for num in sorted(covers, key=int) for col in sorted(covers[num])]
            others = [c for c in df.columns if c not in set(term_id + loc_cols + cover_cols)]
            df = df[[c for c in term_id + loc_cols + cover_cols + others if c in df.columns]]

            for row in ws.iter_rows():
                for cell in row:
                    old_font = cell.font or Font()
                    cell.font = Font(
                        name="Times New Roman",
                        size=12,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color=old_font.color
                    )
                    try:
                        # Try to interpret the cell value as a number
                        float(cell.value)
                        cell.number_format = rupee_format
                    except (TypeError, ValueError):
                        # Skip if it's not a number or cannot be converted to float
                        pass
            ws.sheet_view.zoomScale = 85
            chars_per_line = 85  # Average number of characters per line
            default_row_height = 17  # Default row height in pixels

            # Set all row heights to 17px initially
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = default_row_height

            # Loop through all rows and columns in the worksheet
            for row_num in range(7, ws.max_row + 1):  # Loop through all rows
                max_chars_in_row = 0
                # Loop through all columns in the row
                for col_num in range(1, ws.max_column + 1):  # Loop through all columns
                    cell = ws.cell(row=row_num, column=col_num)
                    if col_num in (4,5,6):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell_data = str(cell.value) if cell.value else ""
                    
                    # Update the max_chars_in_row for the row based on the current cell
                    max_chars_in_row = max(max_chars_in_row, len(cell_data))

                # Estimate the number of lines needed for the row
                num_lines = math.ceil(max_chars_in_row / chars_per_line)

                # Calculate the row height based on the number of lines
                row_height = default_row_height * num_lines

                # Set the row height for this row
                ws.row_dimensions[row_num].height = row_height
            set_dynamic_row_heights(ws, row_height_dict)
            # Set the print area (columns A to F, full used row range)
            last_row = ws.max_row
            ws.print_area = f"A1:F{last_row}"
            # Page layout: Fit columns on one page, allow rows to go to multiple pages
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0  # 0 means "automatic" (multiple pages tall)
            ws.page_setup.fitToWidth = 1             
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[5].height = 17 * line_count
            ws.print_title_rows = '1:7'
            # for row in ws.iter_rows(min_row=1, max_row=6):
            #     for cell in row:
            #         if cell.value == "RISK LOCATION":
            #             target_cell = ws.cell(row=cell.row, column=3)
            #             target_cell.alignment = Alignment(wrap_text=True, vertical='top')

            wb.save(output_file)
            wb.close()

            # Step 4: Drop the original 'Location' column if not needed
            df.drop(columns=['Location'], inplace=True, errors='ignore')            
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Annexure', index=False)
                worksheet = writer.book['Annexure']

                # Auto-adjust column widths
                for col_idx, column_cells in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), start=1):
                    max_length = 0
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2  # Add a little padding
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

prod_rev = pd.read_csv("working_folder/product_input_rev.csv")

# Add empty list-style initialized columns
new_columns = [
    "Strikes", "Strike_SI_Weights",
    "VAR_deductible", "VAR_levels", "VAR_level_size",
    "VAR_basepay_wt", "VAR_payout_type"
]
for col in new_columns:
    prod_rev[col] = "["

# Loop through each row
for data_index, prod_row in prod_rev.iterrows():
    Num_Risk_Units = int(prod_row["Num_Risk_Units"])

    # Prepare parts as lists
    strikes_parts = []
    strike_weights_parts = []
    var_deductible_parts = []
    var_levels_parts = []
    var_level_size_parts = []
    var_basepay_wt_parts = []
    var_payout_type_parts = []

    # Clean input safely
    strike_raw = str(prod_row["Cust_Strikes"])
    payout_raw = str(prod_row["Cust_Payout_Weights"])
    cover_raw = str(prod_row["Covers"])

    strike_chunks = strike_raw.strip('[]').split(';')
    payout_chunks = payout_raw.strip('[]').split(';')
    covers = cover_raw.strip('[]').split(';')

    for RU_index in range(Num_Risk_Units):
        Strikes = [float(w) for w in strike_chunks[RU_index].split('&')]
        Payouts = [float(w) for w in payout_chunks[RU_index].split('&')]
        cover_at_index = covers[RU_index]

        if "VAR" in cover_at_index:
            deductible = Strikes[0]
            num_levels = Strikes[1]
            level_size = Strikes[2] if len(Strikes) >= 3 else 1
            basepay = Payouts[0]
            if len(Payouts) >= 3:
                payout_type = "Discrete" if Payouts[2] == 1 else "Continuous"
            else:
                payout_type = "Discrete"

            var_deductible_parts.append(str(deductible))
            var_levels_parts.append(str(num_levels))
            var_level_size_parts.append(str(level_size))
            var_basepay_wt_parts.append(str(basepay))
            var_payout_type_parts.append(payout_type)

            # Fill NA for strike columns
            strikes_parts.append("NA")
            strike_weights_parts.append("NA")
        else:
            strikes_formatted = '&'.join(str(x) for x in Strikes)
            payouts_formatted = '&'.join(str(x) for x in Payouts)

            strikes_parts.append(strikes_formatted)
            strike_weights_parts.append(payouts_formatted)

            var_deductible_parts.append("NA")
            var_levels_parts.append("NA")
            var_level_size_parts.append("NA")
            var_basepay_wt_parts.append("NA")
            var_payout_type_parts.append("NA")

    # Final assignment without trailing semicolon
    prod_rev.at[data_index, 'Strikes'] = "[" + ';'.join(strikes_parts) + "]"
    prod_rev.at[data_index, 'Strike_SI_Weights'] = "[" + ';'.join(strike_weights_parts) + "]"
    prod_rev.at[data_index, 'VAR_deductible'] = "[" + ';'.join(var_deductible_parts) + "]"
    prod_rev.at[data_index, 'VAR_levels'] = "[" + ';'.join(var_levels_parts) + "]"
    prod_rev.at[data_index, 'VAR_level_size'] = "[" + ';'.join(var_level_size_parts) + "]"
    prod_rev.at[data_index, 'VAR_basepay_wt'] = "[" + ';'.join(var_basepay_wt_parts) + "]"
    prod_rev.at[data_index, 'VAR_payout_type'] = "[" + ';'.join(var_payout_type_parts) + "]"
    
    # Get individual components, replacing with "NA" if missing
    lat_lon_val = str(prod_row["lat_lon"]).strip("[]").replace("&", "_") if pd.notna(prod_row["lat_lon"]) else "NA"
    pincode_val = str(int(prod_row["pincode"])) if pd.notna(prod_row["pincode"]) else "NA"
    sdtname_val = str(prod_row["sdtname"]) if pd.notna(prod_row["sdtname"]) else "NA"
    dtname_val = str(prod_row["dtname"]) if pd.notna(prod_row["dtname"]) else "NA"
    stname_val = str(prod_row["stname"]) if pd.notna(prod_row["stname"]) else "NA"

    # Construct the loc_name
    loc_name = f"[{lat_lon_val};{pincode_val};{sdtname_val};{dtname_val};{stname_val};India]"
    # Assign to dataframe
    prod_rev.at[data_index, "loc_name"] = loc_name
prod_rev = prod_rev.rename(columns={
    "risk": "Risk_Details",
    "Total_SI": "Unit_SI",
    "lat_lon": "Ref_Lat_Lon",
    "Xval": "x_val"
})

prod_rev["Unit_Net_Prem"] = prod_rev["Unit_SI"] * prod_rev["Target_PR"]
prod_rev["Cyclone_Speed_Dist_PayoutWt"] = ""

# Drop unused columns
columns_to_drop = [
    "project_name", "IC_Name", "to_run", "code_mode", "season",
    "min_return_period", "start_year", "end_year",
    "bin_multipliers", "strat_wt",
    "Risk_Weights", "N_Strikes",
    "Cover_deductible", "Cover_levels", "Cover_level_size",
    "Cover_basepay", "Cover_payout_type", "Cover_threshold_min", "Cover_threshold_max",
    "Cust_Strikes", "Cust_Payout_Weights", "Strike_Risk_Weights", "Target_PR", "loc_type", "loc_ref", "pincode", "sdtname", "dtname", "stname", "TLR", "model_version", "run_time"
]
prod_rev = prod_rev.drop(columns=columns_to_drop, errors='ignore')

final_order = [
    "termsheet_id", "Risk_Details", "loc_name", "RSD", "RED",
    "Unit_Type", "Unit_SI", "Unit_Net_Prem", "GST(%)",
    "Num_Risk_Units", "Covers", "Ref_Lat_Lon", "x_val",
    "Data_Sources", "Phase_Dates", "SI_Weights",
    "Strikes", "Strike_SI_Weights",
    "VAR_deductible", "VAR_levels", "VAR_level_size",
    "VAR_basepay_wt", "VAR_payout_type",
    "Cyclone_Speed_Dist_PayoutWt"
]

# Reorder columns in prod_rev
prod_rev = prod_rev[final_order + [col for col in prod_rev.columns if col not in final_order]]

file_path = './working_folder/Termsheet/Termsheet_Uploader.xlsx'
book = load_workbook(file_path)

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    prod_rev.to_excel(writer, sheet_name='ts_uploader', index=False, na_rep="NA")

pdf = sys.argv[1]

convert_termsheet_uploader_to_draft(file_path)



