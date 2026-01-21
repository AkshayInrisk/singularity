import pandas as pd
import numpy as np
import xarray as xr
from datetime import datetime, timedelta
import math
import warnings
import os
import pickle
import shutil
import gc
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter
import Cover_Definitions as cover_definitions
from Cover_Definitions import temp_ds_names
from Cover_Calc import cover_defn, cover_rev_defn
from BackTest import backtest
import subprocess
import hashlib
import json
from git import Repo
from google.cloud import storage
from io import StringIO
import sys

# User Inputs
# User Inputs (read from environment so Cloud Run / app.py controls behavior)
save_data = int(os.environ.get("SAVE_DATA", "1"))
run_analytics = int(os.environ.get("RUN_ANALYTICS", "1"))
save_analytics = int(os.environ.get("SAVE_ANALYTICS", "1"))

termsheet = os.environ.get("TERMSHEET", "No")
pdf = os.environ.get("PDF", "No")


def termsheet_id(df, exclude_columns=None, repo_path='.',
                                  bucket_name="inrisk-gpt-logs",
                                  blob_path="Singularity_usage_logs/usage_logs_singularity.csv"):
    """
    Adds termsheet_id, model_version, and run_time to each row in df.
    Appends df to usage log stored in GCS.
    
    Parameters:
        df (pd.DataFrame): Input dataframe
        exclude_columns (list): Columns to exclude from hashing
        repo_path (str): Path to git repo
        creds_path (str): Path to GCP credentials
        bucket_name (str): GCS bucket
        blob_path (str): Blob path within the bucket
    
    Returns:
        pd.DataFrame: Enriched dataframe with metadata columns
    """

    if exclude_columns is None:
        exclude_columns = []

    # Git version and timestamp
    # model_version = Repo(repo_path).head.commit.hexsha
    model_version = "v1_manual"
    run_time = datetime.now().isoformat(timespec='microseconds')

    # Drop excluded columns for hash computation
    df_for_hash = df.drop(columns=exclude_columns, errors='ignore')

    def generate_row_hash(row_dict):
        row_json = json.dumps(row_dict, sort_keys=True)
        combined = row_json + model_version
        return hashlib.sha256(combined.encode()).hexdigest()

    termsheet_ids = df_for_hash.apply(lambda row: generate_row_hash(row.to_dict()), axis=1)

    # Add metadata columns
    df_new = df.copy()
    df_new['termsheet_id'] = termsheet_ids
    df_new['model_version'] = model_version
    df_new['run_time'] = run_time

    # Setup GCS client (uses Application Default Credentials on Cloud Run)
    if str(os.environ.get('ENABLE_USAGE_LOGS','0')).lower() not in {'1','true','yes'}:
        return df_new
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_path)

    # Append to usage log
    try:
        existing_data = blob.download_as_text()
        existing_df = pd.read_csv(StringIO(existing_data))
        combined_df = pd.concat([existing_df, df_new], ignore_index=True)
    except Exception:
        # If blob doesn't exist or is empty
        combined_df = df_new

    # Upload updated log
    log_buffer = StringIO()
    combined_df.to_csv(log_buffer, index=False)
    blob.upload_from_string(log_buffer.getvalue(), content_type='text/csv')

    return df_new

warnings.filterwarnings("ignore")

from pathlib import Path
import os

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

# Must be set by app.py per request (unique folder)
WORKDIR = Path(os.environ.get("WORKING_DIR", str(BASE_DIR / "working_folder")))
INPUT_DIR = WORKDIR / "input"
OUTPUT_DIR = WORKDIR / "output"

INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Subfolders that pipeline uses
DATA_WORKING_DIR = OUTPUT_DIR / "Data_Working"
TEMP_WORKING_DIR = OUTPUT_DIR / "Temp_Working"
DATA_WORKING_DIR.mkdir(parents=True, exist_ok=True)
TEMP_WORKING_DIR.mkdir(parents=True, exist_ok=True)

# Keep your legacy variables used throughout the file
path = str(OUTPUT_DIR)
temp_path = str(TEMP_WORKING_DIR)



#Read the files and define output locations
block_geo = pd.read_csv(str(DATA_DIR / "block_geo.csv"))
pincode_geo = pd.read_csv(str(DATA_DIR / "pincode_geo.csv"))
prod_file = str(INPUT_DIR / 'product_input.csv')
prod_file_2 = str(OUTPUT_DIR / 'product_input_rev.csv')
base_data_path = str(OUTPUT_DIR / '2. Historical Weather Data.xlsx')
analytics_path = str(OUTPUT_DIR / '3. Product Analytics.xlsx')
claims_path = str(OUTPUT_DIR / '4. Policy_Monitoring.xlsx')

products = pd.read_csv(prod_file)
products_2 = termsheet_id(products, exclude_columns=["project_name","IC_Name","to_run","Unit_Type", "GST(%)", "risk"])
products_2['Cust_Strikes'] = pd.Series(dtype=object)
products_2['Cust_Payout_Weights'] = pd.Series(dtype=object)

# Decide whether to compute data phase or load cached artifacts (per request)
risk_datas_file = Path(path) / "Data_Working" / "Risk_Datas.parquet"

if risk_datas_file.exists():
    run_data = 0
    save_data_dump = 0
else:
    run_data = 1
    save_data_dump = 1

if run_data == 1:
    data_start_time = datetime.now()
    ### DATA PHASE COMMENCES ###

    data_records = []
    start_year = 9999
    end_year = 9999

    All_Data_Sources = products['Data_Sources'].apply(lambda x: x.strip('[]').split(';'))
    All_Unique_Data_Sources = set([item for sublist in All_Data_Sources for item in sublist])

    State_Names, Dist_Names, Block_Names, Pincodes, Locations = {}, {}, {}, {}, {}

    for data_index, data_row in products.iterrows():
        # products = pd.read_csv(prod_file)
        # data_index = 0
        # data_row = products.loc[data_index]

        ## Pricing Model Selection ##
        code_mode = data_row["code_mode"] #'reverse', 'custom', 'claims'

        RSD = pd.to_datetime(data_row["RSD"], format="%d-%m-%Y")
        RED = pd.to_datetime(data_row["RED"], format="%d-%m-%Y")

        ## Historical Analysis Assumptions ##
        if code_mode == 'claims':
            start_year = min(RSD.year, start_year)
            end_year = start_year + (RED.year - RSD.year)
        elif code_mode in ('reverse', 'custom'):
            start_year = min(start_year, int(data_row["start_year"]))
            end_year = min(end_year, int(data_row["end_year"]))

        loc_type = data_row["loc_type"] # 'Pincode' or 'Block'
        loc_ref = data_row["loc_ref"] # "Name" or "Coordinates"

        State_Name = data_row["stname"]
        State_Names[data_index] = State_Name
        Dist_Name = data_row["dtname"]
        Dist_Names[data_index] = Dist_Name
        Block_Name = data_row["sdtname"]
        Block_Names[data_index] = Block_Name
        Pincode = data_row["pincode"]
        Pincodes[data_index] = Pincode
        
        if loc_ref == "Name":
            if loc_type == "Pincode":
                geo_df = pincode_geo[pincode_geo['pincode'] == Pincode]
            elif loc_type == "Block":
                geo_df = block_geo[
                    (block_geo["stname"].str.strip().str.upper() == State_Name.strip().upper()) &
                    (block_geo["dtname"].str.strip().str.upper() == Dist_Name.strip().upper()) &
                    (block_geo["sdtname"].str.strip().str.upper() == Block_Name.strip().upper())
                ]

        Num_Risk_Units = int(data_row["Num_Risk_Units"])

        for RU_index in range(Num_Risk_Units):
            # RU_index = 0
            source = data_row['Data_Sources'].strip('[]').split(";")[RU_index]

            if "IMD_RAIN" in source:
                data_type = "IMD_RF"
                lat_col, lon_col = "lat_0_25", "lon_0_25"
            elif "IMD_TMAX" in source or "IMD_TMIN" in source or "IMD_TAVG" in source or "IMD_TRANGE" in source:
                data_type = "IMD_TEMP"
                if loc_type == "Pincode" and 'IMD_TAVG' in source:
                    lat_col, lon_col = "lat_0_5_avg", "lon_0_5_avg"
                else:
                    lat_col, lon_col = "lat_0_5", "lon_0_5"
            elif "ERA5L_" in source:
                data_type = "ERA5_LAND"
                lat_col, lon_col = "lat_0_1", "lon_0_1"
            elif "ERA5_" in source:
                data_type = "ERA5"
                lat_col, lon_col = "lat_0_25", "lon_0_25"
            else:
                continue  # Skip unknown data sources

            if loc_ref == "Name":
                lat, lon = geo_df[lat_col].values[0], geo_df[lon_col].values[0]                
            elif loc_ref == "Coordinates":
                lat, lon = data_row['lat_lon'].strip("[]").split(";")[RU_index].split("&")
                
            Locations[f"{data_index}.{RU_index}"] = [float(lat), float(lon)]
            data_records.append((data_type, float(lat), float(lon), source))

    # Create DataFrame with unique values
    df_unique = pd.DataFrame(data_records, columns=["Data_Source_Type", "latitude", "longitude", "Data_Source_Name"])
    df_unique = df_unique.drop_duplicates().reset_index(drop=True)
    df_unique = df_unique.sort_values(by=["Data_Source_Type", "Data_Source_Name", "latitude", "longitude"]).reset_index(drop=True)

    print("Data Indexes Created.")

    combined_data = pd.DataFrame()

    if len(df_unique[df_unique["Data_Source_Type"] == "ERA5"]) > 0:
        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        era5_index = pd.read_csv(str(DATA_DIR / "era5_index.csv"))
        bq_client = bigquery.Client()
        bqstorage_client = bigquery_storage.BigQueryReadClient()

        era5_data = df_unique[df_unique["Data_Source_Type"] == "ERA5"][["Data_Source_Name"]].drop_duplicates().reset_index()
        era5_points = df_unique[df_unique["Data_Source_Type"] == "ERA5"][["latitude", "longitude"]].drop_duplicates()
        era5_points = pd.merge(era5_points, era5_index, on=["latitude", "longitude"], how="left")
        
        indices = sorted(np.array(era5_points["index"].unique()))
        indices_str = ", ".join(map(str, indices))

        era5_variable_dict = {
            # Precipitation & runoff
            "ERA5_RAIN": "tp",
            "ERA5_RUNOFF": "ro",
            "ERA5_SURFACE_RUNOFF": "sro",

            # Temperature (2m)
            "ERA5_TMAX": "t2m_max",
            "ERA5_TAVG": "t2m_mean",
            "ERA5_TMIN": "t2m_min",

            # Dew point temperature (2m)
            "ERA5_D2MMAX": "d2m_max",
            "ERA5_D2MAVG": "d2m_mean",
            "ERA5_D2MMIN": "d2m_min",

            # Relative humidity
            "ERA5_RHMAX": "rh_max",
            "ERA5_RHAVG": "rh_mean",
            "ERA5_RHMIN": "rh_min",

            # Heat Index
            "ERA5_HIMAX": "hi_max",
            "ERA5_HIAVG": "hi_mean",
            "ERA5_HIMIN": "hi_min",

            # Temperature Humidity Index
            "ERA5_THIMAX": "thi_max",
            "ERA5_THIAVG": "thi_mean",
            "ERA5_THIMIN": "thi_min",

            # Wind speed (10m derived)
            "ERA5_WSMAX": "ws_max",
            "ERA5_WSAVG": "ws_mean",
            "ERA5_WSMIN": "ws_min",

            # Instantaneous wind gust (10m)
            "ERA5_I10FGMAX": "i10fg_max",
            "ERA5_I10FGAVG": "i10fg_mean",
            "ERA5_I10FGMIN": "i10fg_min",

            # Wind components (10m)
            "ERA5_U10MAX": "u10_max",
            "ERA5_U10AVG": "u10_mean",
            "ERA5_U10MIN": "u10_min",
            "ERA5_V10MAX": "v10_max",
            "ERA5_V10AVG": "v10_mean",
            "ERA5_V10MIN": "v10_min",

            # Wind components (100m)
            "ERA5_U100MAX": "u100_max",
            "ERA5_U100AVG": "u100_mean",
            "ERA5_U100MIN": "u100_min",
            "ERA5_V100MAX": "v100_max",
            "ERA5_V100AVG": "v100_mean",
            "ERA5_V100MIN": "v100_min",

            # Mean sea level pressure
            "ERA5_MSLMAX": "msl_max",
            "ERA5_MSLAVG": "msl_mean",
            "ERA5_MSLMIN": "msl_min",

            # Surface pressure
            "ERA5_SPMAX": "sp_max",
            "ERA5_SPAVG": "sp_mean",
            "ERA5_SPMIN": "sp_min",

            # Sea surface temperature
            "ERA5_SSTMAX": "sst_max",
            "ERA5_SSTAVG": "sst_mean",
            "ERA5_SSTMIN": "sst_min",

            # Radiation
            "ERA5_SSRDNET": "ssrd",

            # Evaporation
            "ERA5_EVAP": "e",
            "ERA5_PEV": "pev",

            # Vegetation / soil
            "ERA5_LAILVAVG": "lai_lv_mean",
            "ERA5_LAIHVAVG": "lai_hv_mean",
            "ERA5_SOILM1": "swvl1_mean",
            "ERA5_SOILM2": "swvl2_mean",
            "ERA5_SOILM3": "swvl3_mean",
            "ERA5_SOILM4": "swvl4_mean"
        }


        reverse_era5_variable_dict = {v: k for k, v in era5_variable_dict.items()}

        # Cumulative variables live ONLY in era5_cum table
        CUM_VARS = {"e", "pev","ro", "sro", "ssrd", "tp"}

        era5_inst_vars = set()
        era5_cum_vars = set()

        for name in era5_data["Data_Source_Name"]:

            # Composite variable → expand into two base ERA5 variables
            if name.count('_') == 4:
                ds_components = name.split("_")[1:3]  # take first two only

                for comp in ds_components:
                    era5_name = "ERA5_" + comp
                    col = era5_variable_dict.get(era5_name, "")
                    if not col:
                        continue

                    short_var = col.lower()
                    if short_var in CUM_VARS:
                        era5_cum_vars.add(col)
                    else:
                        era5_inst_vars.add(col)

            # Normal variable
            else:
                col = era5_variable_dict.get(name, "")
                if not col:
                    continue

                short_var = col.lower()
                if short_var in CUM_VARS:
                    era5_cum_vars.add(col)
                else:
                    era5_inst_vars.add(col)

        # Final SQL-ready strings
        era5_inst_variables = ", " + ", ".join(sorted(era5_inst_vars)) if era5_inst_vars else ""
        era5_cum_variables  = ", " + ", ".join(sorted(era5_cum_vars))  if era5_cum_vars  else ""
        # Query instantaneous variables
        query_inst = f"""
            SELECT valid_time, latitude, longitude{era5_inst_variables}
            FROM `firm-aviary-409110.inrisk_db.era5_inst`
            WHERE `index` IN ({indices_str}) 
            AND EXTRACT(YEAR FROM valid_time) >= {start_year} 
            AND EXTRACT(YEAR FROM valid_time) <= {end_year}
        """
        query_job = bq_client.query(query_inst)
        print("Extracting ERA5 Data...")
        era5_df = query_job.to_dataframe(bqstorage_client=bqstorage_client)
        print("ERA5 Instantaneous Data Extracted")

        # Query cumulative variables (tp)
        query_cum = f"""
            SELECT valid_time, latitude, longitude{era5_cum_variables}
            FROM `firm-aviary-409110.inrisk_db.era5_cum`
            WHERE `index` IN ({indices_str}) 
            AND EXTRACT(YEAR FROM valid_time) >= {start_year} 
            AND EXTRACT(YEAR FROM valid_time) <= {end_year}
        """
        query_job_cum = bq_client.query(query_cum)
        era5_cum_df = query_job_cum.to_dataframe(bqstorage_client=bqstorage_client)
        print("ERA5 Cumulative Data Extracted")

        # Merge cumulative variables into era5_df
        era5_df = era5_df.merge(
            era5_cum_df, 
            on=['valid_time', 'latitude', 'longitude'], 
            how='left'
        )
        print("Merged ERA5 Instantaneous and Cumulative Data")

        era5_df = pd.merge(era5_df, era5_points, on=["latitude", "longitude"], how="inner")
        era5_df.drop(columns=["index"], inplace=True)
        era5_df["latitude"] = era5_df["latitude"].astype(float)
        era5_df["longitude"] = era5_df["longitude"].astype(float)
        era5_df["valid_date"] = pd.to_datetime(era5_df["valid_time"])
        era5_df.drop(columns=["valid_time"], inplace=True)
        if "tp" in era5_df.columns:
            era5_df["tp"] = era5_df["tp"] * 1000
        if "ssr" in era5_df.columns:
            era5_df["ssr"] = era5_df["ssr"] / (3.6 * 10**6)
        if "ssrd" in era5_df.columns:
            era5_df["ssrd"] = era5_df["ssrd"] / (3.6 * 10**6)
        era5_df = era5_df.sort_values(by="valid_date")
        era5_df.rename(columns=reverse_era5_variable_dict, inplace=True)
        era5_df.rename(columns={"valid_date":"time", "latitude":"lat", "longitude":"lon"}, inplace=True)

        if name.count('_') > 1:
            # E.g. ERA5_TMAX_ERA5_WSMIN_LTE_15

            ds_components = name.split("_")
            name_1 = ds_components[0] + "_" + ds_components[1]

            name_2 = ds_components[2] + "_" + ds_components[3]
            gt_type = ds_components[4]
            gt_val = float(ds_components[5])

            if gt_type == "GTE":
                era5_df[name] = era5_df.apply(lambda row: -1 if row[name_2] < gt_val else row[name_1], axis=1)
            else:
                era5_df[name] = era5_df.apply(lambda row: -1 if row[name_2] > gt_val else row[name_1], axis=1)

        era5_melted = era5_df.melt(id_vars=["time", "lat", "lon"], var_name="data", value_name="parameter")
        combined_data = pd.concat([combined_data, era5_melted], ignore_index=True)

    if len(df_unique[df_unique["Data_Source_Type"] == "ERA5_LAND"]) > 0:
        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        era5_index = pd.read_csv(str(DATA_DIR / "era5_land_index.csv"))
        bq_client = bigquery.Client()
        bqstorage_client = bigquery_storage.BigQueryReadClient()

        era5_land_data = df_unique[df_unique["Data_Source_Type"] == "ERA5_LAND"][["Data_Source_Name"]].drop_duplicates().reset_index()
        era5_land_points = df_unique[df_unique["Data_Source_Type"] == "ERA5_LAND"][["latitude", "longitude"]].drop_duplicates()
        era5_land_points = pd.merge(era5_land_points, era5_index, on=["latitude", "longitude"], how="left")
        
        indices = sorted(np.array(era5_land_points["index"].unique()))
        indices_str = ", ".join(map(str, indices))

        era5_variable_dict = {

            # Precipitation & runoff
            "ERA5L_RAIN": "tp",
            "ERA5L_RUNOFF": "ro",
            "ERA5L_SURFACE_RUNOFF": "sro",

            # 2m Air Temperature
            "ERA5L_TMAX": "t2m_max",
            "ERA5L_TAVG": "t2m_mean",
            "ERA5L_TMIN": "t2m_min",

            # 2m Dew Point Temperature
            "ERA5L_D2MMAX": "d2m_max",
            "ERA5L_D2MAVG": "d2m_mean",
            "ERA5L_D2MMIN": "d2m_min",

            # Relative Humidity
            "ERA5L_RHMAX": "rh_max",
            "ERA5L_RHAVG": "rh_mean",
            "ERA5L_RHMIN": "rh_min",

            # Heat Index
            "ERA5L_HIMAX": "hi_max",
            "ERA5L_HIAVG": "hi_mean",
            "ERA5L_HIMIN": "hi_min",

            # Temperature Humidity Index
            "ERA5L_THIMAX": "thi_max",
            "ERA5L_THIAVG": "thi_mean",
            "ERA5L_THIMIN": "thi_min",

            # Skin Temperature
            "ERA5L_SKTMX": "skt_max",
            "ERA5L_SKTMN": "skt_min",
            "ERA5L_SKTAVG": "skt_mean",

            # Surface Pressure
            "ERA5L_SPMAX": "sp_max",
            "ERA5L_SPAVG": "sp_mean",
            "ERA5L_SPMIN": "sp_min",

            # Solar Radiation
            "ERA5L_SSRNET": "ssr",
            "ERA5L_SSRDNET": "ssrd",

            # Soil Temperature (levels 1–4)
            "ERA5L_STL1MAX": "stl1_max",
            "ERA5L_STL1AVG": "stl1_mean",
            "ERA5L_STL1MIN": "stl1_min",

            "ERA5L_STL2MAX": "stl2_max",
            "ERA5L_STL2AVG": "stl2_mean",
            "ERA5L_STL2MIN": "stl2_min",

            "ERA5L_STL3MAX": "stl3_max",
            "ERA5L_STL3AVG": "stl3_mean",
            "ERA5L_STL3MIN": "stl3_min",

            "ERA5L_STL4MAX": "stl4_max",
            "ERA5L_STL4AVG": "stl4_mean",
            "ERA5L_STL4MIN": "stl4_min",

            # Soil Moisture (levels 1–4)
            "ERA5L_SOILM1MAX": "swvl1_max",
            "ERA5L_SOILM1AVG": "swvl1_mean",
            "ERA5L_SOILM1MIN": "swvl1_min",

            "ERA5L_SOILM2MAX": "swvl2_max",
            "ERA5L_SOILM2AVG": "swvl2_mean",
            "ERA5L_SOILM2MIN": "swvl2_min",

            "ERA5L_SOILM3MAX": "swvl3_max",
            "ERA5L_SOILM3AVG": "swvl3_mean",
            "ERA5L_SOILM3MIN": "swvl3_min",

            "ERA5L_SOILM4MAX": "swvl4_max",
            "ERA5L_SOILM4AVG": "swvl4_mean",
            "ERA5L_SOILM4MIN": "swvl4_min",

            # Leaf Area Index
            "ERA5L_LAIHVMAX": "lai_hv_max",
            "ERA5L_LAIHVAVG": "lai_hv_mean",
            "ERA5L_LAIHVMIN": "lai_hv_min",

            "ERA5L_LAILVMAX": "lai_lv_max",
            "ERA5L_LAILVAVG": "lai_lv_mean",
            "ERA5L_LAILVMIN": "lai_lv_min",

            # Wind (10m)
            "ERA5L_U10MAX": "u10_max",
            "ERA5L_U10AVG": "u10_mean",
            "ERA5L_U10MIN": "u10_min",

            "ERA5L_V10MAX": "v10_max",
            "ERA5L_V10AVG": "v10_mean",
            "ERA5L_V10MIN": "v10_min",

            "ERA5L_WSMAX": "ws_max",
            "ERA5L_WSAVG": "ws_mean",
            "ERA5L_WSMIN": "ws_min",

            # Shortwave Radiation (Cloud-modulated)
            "ERA5L_SRCMAX": "src_max",
            "ERA5L_SRCAVG": "src_mean",
            "ERA5L_SRCMIN": "src_min"
        }

        reverse_era5_variable_dict = {v: k for k, v in era5_variable_dict.items()}

        # Cumulative variables live ONLY in era5_land_cum table
        CUM_VARS = {"ro", "sro", "ssr", "ssrd", "tp"}

        era5_inst_vars = set()
        era5_cum_vars = set()

        for name in era5_land_data["Data_Source_Name"]:

            # Composite variable → expand into two base ERA5L variables
            if name.count('_') == 4:
                ds_components = name.split("_")[1:3]  # take first two only

                for comp in ds_components:
                    era5_name = "ERA5L_" + comp
                    col = era5_variable_dict.get(era5_name, "")
                    if not col:
                        continue

                    short_var = col.lower()
                    if short_var in CUM_VARS:
                        era5_cum_vars.add(col)
                    else:
                        era5_inst_vars.add(col)

            # Normal variable
            else:
                col = era5_variable_dict.get(name, "")
                if not col:
                    continue

                short_var = col.lower()
                if short_var in CUM_VARS:
                    era5_cum_vars.add(col)
                else:
                    era5_inst_vars.add(col)

        # Final SQL-ready strings
        era5_inst_variables = ", " + ", ".join(sorted(era5_inst_vars)) if era5_inst_vars else ""
        era5_cum_variables  = ", " + ", ".join(sorted(era5_cum_vars))  if era5_cum_vars  else ""

        # Query instantaneous variables
        query_inst = f"""
            SELECT valid_time, latitude, longitude{era5_inst_variables}
            FROM `firm-aviary-409110.inrisk_db.era5_land_inst`
            WHERE `index` IN ({indices_str}) 
            AND EXTRACT(YEAR FROM valid_time) >= {start_year} 
            AND EXTRACT(YEAR FROM valid_time) <= {end_year}
        """
        query_job = bq_client.query(query_inst)
        print("Extracting ERA5-Land Data...")
        era5_land_df = query_job.to_dataframe(bqstorage_client=bqstorage_client)
        print("ERA5-Land Instantaneous Data Extracted")

        # Query cumulative variables (tp)
        query_cum = f"""
            SELECT valid_time, latitude, longitude{era5_cum_variables}
            FROM `firm-aviary-409110.inrisk_db.era5_land_cum`
            WHERE `index` IN ({indices_str}) 
            AND EXTRACT(YEAR FROM valid_time) >= {start_year} 
            AND EXTRACT(YEAR FROM valid_time) <= {end_year}
        """
        query_job_cum = bq_client.query(query_cum)
        era5_land_cum_df = query_job_cum.to_dataframe(bqstorage_client=bqstorage_client)
        print("ERA5-Land Cumulative Data Extracted")

        # Merge cumulative variables into era5_land_df
        era5_land_df = era5_land_df.merge(
            era5_land_cum_df, 
            on=['valid_time', 'latitude', 'longitude'], 
            how='left'
        )
        print("Merged ERA5-Land Instantaneous and Cumulative Data")


        era5_land_df = pd.merge(era5_land_df, era5_land_points, on=["latitude", "longitude"], how="inner")
        era5_land_df.drop(columns=["index"], inplace=True)
        era5_land_df["latitude"] = era5_land_df["latitude"].astype(float)
        era5_land_df["longitude"] = era5_land_df["longitude"].astype(float)
        era5_land_df["valid_date"] = pd.to_datetime(era5_land_df["valid_time"])
        era5_land_df.drop(columns=["valid_time"], inplace=True)
        if "tp" in era5_land_df.columns:
            era5_land_df["tp"] = era5_land_df["tp"] * 1000
        # if "rh_max" in era5_land_df.columns:
        #     era5_land_df["rh_max"] = era5_land_df["rh_max"] * 100
        # if "rh_min" in era5_land_df.columns:
        #     era5_land_df["rh_min"] = era5_land_df["rh_min"] * 100
        # if "rh_mean" in era5_land_df.columns:
        #     era5_land_df["rh_mean"] = era5_land_df["rh_mean"] * 100
        if "ssr" in era5_land_df.columns:
            era5_land_df["ssr"] = era5_land_df["ssr"] / (3.6 * 10**6)
        if "ssrd" in era5_land_df.columns:
            era5_land_df["ssrd"] = era5_land_df["ssrd"] / (3.6 * 10**6)
        era5_land_df = era5_land_df.sort_values(by="valid_date")
        era5_land_df.rename(columns=reverse_era5_variable_dict, inplace=True)
        era5_land_df.rename(columns={"valid_date":"time", "latitude":"lat", "longitude":"lon"}, inplace=True)
        for name in era5_land_data["Data_Source_Name"]:
            # name = era5_land_data["Data_Source_Name"][0]
            if name.count('_') == 4:
                ds_components = name.split("_")[1:]
                name_1 = "ERA5L_" + ds_components[0]
                name_2 = "ERA5L_" + ds_components[1]

                gt_type = ds_components[2]
                gt_val = float(ds_components[3])

                if gt_type == "GTE":
                    era5_land_df[name] = era5_land_df.apply(lambda row: -1 if row[name_2] < gt_val else row[name_1], axis=1)
                else:
                    era5_land_df[name] = era5_land_df.apply(lambda row: -1 if row[name_2] > gt_val else row[name_1], axis=1)

        era5_land_melted = era5_land_df.melt(id_vars=["time", "lat", "lon"], var_name="data", value_name="parameter")
        combined_data = pd.concat([combined_data, era5_land_melted], ignore_index=True)

    if len(df_unique[df_unique["Data_Source_Type"] == "IMD_RF"]) > 0:
        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        bq_client = bigquery.Client()
        bqstorage_client = bigquery_storage.BigQueryReadClient()

        imd_rf_data = df_unique[df_unique["Data_Source_Type"] == "IMD_RF"][["Data_Source_Name"]].drop_duplicates().reset_index()
        imd_rf_points = df_unique[df_unique["Data_Source_Type"] == "IMD_RF"][["latitude", "longitude"]].drop_duplicates().reset_index(drop=True)
        query_1 = f"""SELECT lat as latitude, lon as longitude, index FROM `firm-aviary-409110.inrisk_db.imd_rain_index`"""
        query_job = bq_client.query(query_1)

        imd_rf_index = query_job.to_dataframe(bqstorage_client=bqstorage_client)

        imd_rf_points = pd.merge(imd_rf_points, imd_rf_index, on=["latitude", "longitude"], how="left")
        indices = sorted(np.array(imd_rf_points["index"].unique()))
        indices_str = ", ".join(map(str, indices))

        query_2 = f"""SELECT time, lat, lon, rain FROM `firm-aviary-409110.inrisk_db.imd_rain_df` WHERE `index` IN ({indices_str}) and EXTRACT(YEAR FROM time) >= {start_year} and EXTRACT(YEAR FROM time) <= {end_year} """

        query_job = bq_client.query(query_2)

        print("Extracting IMD-RAIN Data...")
        imd_rf_df = query_job.to_dataframe(bqstorage_client=bqstorage_client)
        print("Data Extracted")
                
        imd_rf_df["lat"] = imd_rf_df["lat"].astype(float)
        imd_rf_df["lon"] = imd_rf_df["lon"].astype(float)
        imd_rf_df["time"] = pd.to_datetime(imd_rf_df["time"])
        imd_rf_df.sort_values("time", inplace=True)
        imd_rf_df.reset_index(drop=True, inplace=True)        
        imd_rf_df.rename(columns={"rain": "IMD_RAIN"}, inplace=True)
        imd_rf_melted = imd_rf_df.melt(id_vars=["time", "lat", "lon"], var_name="data", value_name="parameter")
        imd_rf_melted = imd_rf_melted[imd_rf_melted['parameter'] >= 0]
        combined_data = pd.concat([combined_data, imd_rf_melted], ignore_index=True)

    if len(df_unique[df_unique["Data_Source_Type"] == "IMD_TEMP"]) > 0:
        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        bq_client = bigquery.Client()
        bqstorage_client = bigquery_storage.BigQueryReadClient()

        imd_temp_data = df_unique[df_unique["Data_Source_Type"] == "IMD_TEMP"][["Data_Source_Name"]].drop_duplicates().reset_index()
        imd_temp_points = df_unique[df_unique["Data_Source_Type"] == "IMD_TEMP"][["latitude", "longitude"]].drop_duplicates().reset_index(drop=True)
        query_1 = f"""SELECT lat as latitude, lon as longitude, index FROM `firm-aviary-409110.inrisk_db.imd_temp_index`"""
        query_job = bq_client.query(query_1)

        imd_temp_index = query_job.to_dataframe(bqstorage_client=bqstorage_client)
        imd_temp_points = pd.merge(imd_temp_points, imd_temp_index, on=["latitude", "longitude"], how="left")
        indices = sorted(np.array(imd_temp_points["index"].unique()))
        indices_str = ", ".join(map(str, indices))

        imd_temp_variable_dict = {
            "IMD_TMAX": "tmax",
            "IMD_TMIN": "tmin",
            # "IMD_TAVG": "tavg",
            "IMD_TRANGE": "trange"
        }
        reverse_imd_temp_variable_dict = {v: k for k, v in imd_temp_variable_dict.items()}
        imd_temp_variable_set  = set()
        compute_tavg = False
        for name in imd_temp_data["Data_Source_Name"]:
            if name == "IMD_TAVG":
                compute_tavg = True
                imd_temp_variable_set .update(["tmax", "tmin"])
            elif name in imd_temp_variable_dict:
                imd_temp_variable_set .add(imd_temp_variable_dict[name])

        # Convert set to string suitable for SQL query
        imd_temp_variables = ", " + ", ".join(sorted(imd_temp_variable_set ))
        query_2 = f"""SELECT time, lat, lon{imd_temp_variables} FROM `firm-aviary-409110.inrisk_db.imd_temp_df` WHERE `index` IN ({indices_str}) and EXTRACT(YEAR FROM time) >= {start_year} and EXTRACT(YEAR FROM time) <= {end_year} """

        query_job = bq_client.query(query_2)
        print("Extracting IMD-TEMP Data...")
        imd_temp_df = query_job.to_dataframe(bqstorage_client=bqstorage_client)
        print("Data Extracted")
        
        imd_temp_df.rename(columns=reverse_imd_temp_variable_dict, inplace=True)

        if compute_tavg:
            imd_temp_df["IMD_TAVG"] = (imd_temp_df["IMD_TMAX"] + imd_temp_df["IMD_TMIN"]) / 2
            protected_cols = {"lat", "lon", "time"}
            candidates_to_drop = {"IMD_TMAX", "IMD_TMIN"}
            unused_columns = candidates_to_drop.difference(set(imd_temp_data["Data_Source_Name"]))
            imd_temp_df.drop(columns=[col for col in unused_columns if col in imd_temp_df.columns and col not in protected_cols], inplace=True)
        
        imd_temp_df["lat"] = imd_temp_df["lat"].astype(float)
        imd_temp_df["lon"] = imd_temp_df["lon"].astype(float)
        imd_temp_df["time"] = pd.to_datetime(imd_temp_df["time"])
        imd_temp_df.sort_values("time", inplace=True)
        imd_temp_df.reset_index(drop=True, inplace=True)
        imd_temp_melted = imd_temp_df.melt(id_vars=["time", "lat", "lon"], var_name="data", value_name="parameter")
        imd_temp_melted = imd_temp_melted[imd_temp_melted['parameter'] < 99]
        combined_data = pd.concat([combined_data, imd_temp_melted], ignore_index=True)

    Risk_Datas = pd.DataFrame()
    
    for yr in range(start_year, end_year+1):
        # yr=1975
        start_date = RSD.replace(year=yr)

        day_adjust = 1 if (RED.month == 2 and RED.day == 29) else 0
        year_adjust = RED.year - RSD.year
        end_date = (RED - timedelta(days=day_adjust)).replace(year=(yr + year_adjust))

        df_filtered = combined_data[(combined_data['time'] >= start_date) & (combined_data['time'] <= end_date)].copy()
        df_filtered['year'] = start_date.year
        Risk_Datas = pd.concat([Risk_Datas, df_filtered], ignore_index=True, sort=False)
    
    Risk_Datas["parameter"] = Risk_Datas["parameter"].round(4)
    Risk_Datas["data"] = Risk_Datas["data"].astype("category")
    Risk_Datas = Risk_Datas.set_index(["lat", "lon", "data"])
    Risk_Datas = Risk_Datas.sort_index()
    print("Data Dump Done.")
    
    if (save_data == 1) or (code_mode == "claims"):
        Data_pivots = pd.DataFrame()
    
        for data_index, data_row in products.iterrows():
            # data_index = 0
            # data_row = products.loc[data_index]
            if data_row["to_run"] == 0:
                continue

            Num_Risk_Units = int(data_row["Num_Risk_Units"])

            data_sources = data_row['Data_Sources'].strip('[]').split(";")
            Risk_Data = pd.DataFrame()
            for RU_index in range(Num_Risk_Units):
                # RU_index = 0
                source = data_sources[RU_index]

                if source.count('_') == 4:
                    ds_components = source.split("_")
                    name_1 = ds_components[0] + "_" + ds_components[1]
                    name_2 = ds_components[0] + "_" + ds_components[2]

                    Risk_Data = pd.concat([Risk_Data, Risk_Datas.loc[Locations[f"{data_index}.{RU_index}"][0], 
                                        Locations[f"{data_index}.{RU_index}"][1], name_1].reset_index()], ignore_index=True, sort=False)
                    Risk_Data = pd.concat([Risk_Data, Risk_Datas.loc[Locations[f"{data_index}.{RU_index}"][0], 
                                        Locations[f"{data_index}.{RU_index}"][1], name_2].reset_index()], ignore_index=True, sort=False)
                    Risk_Data = pd.concat([Risk_Data, Risk_Datas.loc[Locations[f"{data_index}.{RU_index}"][0], 
                                        Locations[f"{data_index}.{RU_index}"][1], source].reset_index()], ignore_index=True, sort=False)

                else:
                    Risk_Data = pd.concat([Risk_Data, Risk_Datas.loc[Locations[f"{data_index}.{RU_index}"][0], 
                                        Locations[f"{data_index}.{RU_index}"][1], source].reset_index()], ignore_index=True, sort=False)

            Data_pivot = Risk_Data.pivot_table(index=['time', 'lat', 'lon', 'year'], columns='data', values='parameter', aggfunc='first').reset_index()
            Data_pivot['State'] = State_Names[data_index].title() if isinstance(State_Names[data_index], str) else 'NA'
            Data_pivot['District'] = Dist_Names[data_index].title() if isinstance(Dist_Names[data_index], str) else 'NA'
            Data_pivot['Block'] = Block_Names[data_index].title() if isinstance(Block_Names[data_index], str) else 'NA'
            Data_pivot['Pincode'] = Pincodes[data_index]
            Data_pivot = Data_pivot.rename(columns={'time':'Date', 'lat':'Ref_Lat', 'lon':'Ref_Lon'})

            desired_order = ['State', 'District', 'Block', 'Pincode', 'Ref_Lat', 'Ref_Lon', 'Date', 'year']
            remaining_columns = [col for col in Data_pivot.columns if col not in desired_order]
            final_order = desired_order + remaining_columns

            Data_pivot = Data_pivot[final_order]
            Data_pivots = pd.concat([Data_pivots, Data_pivot], ignore_index=True)
            Data_pivots['Date'] = pd.to_datetime(Data_pivots['Date'], errors='coerce')
            Data_pivots.fillna('NA', inplace=True)
            Data_pivots.drop_duplicates(inplace=True)

        Data_pivots.replace("NA", np.nan, inplace=True)
        cols_to_fill = ['State', 'District', 'Block', 'Pincode']
        Data_pivots[cols_to_fill] = Data_pivots[cols_to_fill].fillna('NA')

        id_cols  = ['State','District','Block','Pincode','Date','year']
        ref_cols = ['Ref_Lat','Ref_Lon']
        all_data = [c for c in Data_pivots.columns if c not in id_cols + ref_cols]

        era5_cols      = [c for c in all_data if c.startswith('ERA5_')]
        era5L_cols      = [c for c in all_data if c.startswith('ERA5L_')]
        imd_rain_col   = ['IMD_RAIN'] if 'IMD_RAIN' in all_data else []
        imd_temp_cols  = [c for c in all_data if c.startswith('IMD_T')]
        other_cols     = [c for c in all_data 
                        if c not in era5_cols + imd_rain_col + imd_temp_cols]

        if era5_cols:
            Data_pivots['ERA5_Ref_Lat']    = np.where(Data_pivots[era5_cols].notna().any(axis=1),
                                                    Data_pivots['Ref_Lat'], np.nan)
            Data_pivots['ERA5_Ref_Lon']    = np.where(Data_pivots[era5_cols].notna().any(axis=1),
                                                    Data_pivots['Ref_Lon'], np.nan)
        if era5L_cols:
            Data_pivots['ERA5L_Ref_Lat']    = np.where(Data_pivots[era5L_cols].notna().any(axis=1),
                                                    Data_pivots['Ref_Lat'], np.nan)
            Data_pivots['ERA5L_Ref_Lon']    = np.where(Data_pivots[era5L_cols].notna().any(axis=1),
                                                    Data_pivots['Ref_Lon'], np.nan)

        if imd_rain_col:
            Data_pivots['IMD_RAIN_Ref_Lat'] = np.where(
                Data_pivots['IMD_RAIN'].notna(), Data_pivots['Ref_Lat'], np.nan)
            Data_pivots['IMD_RAIN_Ref_Lon'] = np.where(
                Data_pivots['IMD_RAIN'].notna(), Data_pivots['Ref_Lon'], np.nan)

        if imd_temp_cols:
            Data_pivots['IMD_TEMP_Ref_Lat'] = np.where(
                Data_pivots[imd_temp_cols].notna().any(axis=1), Data_pivots['Ref_Lat'], np.nan)
            Data_pivots['IMD_TEMP_Ref_Lon'] = np.where(
                Data_pivots[imd_temp_cols].notna().any(axis=1), Data_pivots['Ref_Lon'], np.nan)

        for col in other_cols:
            Data_pivots[f'{col}_Ref_Lat'] = np.where(
                Data_pivots[col].notna(), Data_pivots['Ref_Lat'], np.nan)
            Data_pivots[f'{col}_Ref_Lon'] = np.where(
                Data_pivots[col].notna(), Data_pivots['Ref_Lon'], np.nan)

        Data_pivots = Data_pivots.drop(ref_cols, axis=1)

        # 5. Build your agg‐dict
        agg = {c: 'first' for c in all_data}
        if era5_cols:
            agg.update({
                'ERA5_Ref_Lat':    'first',
                'ERA5_Ref_Lon':    'first',
            })
        if era5L_cols:
            agg.update({
                'ERA5L_Ref_Lat':    'first',
                'ERA5L_Ref_Lon':    'first',
            })
        if imd_rain_col:
            agg.update({'IMD_RAIN_Ref_Lat': 'first',
                        'IMD_RAIN_Ref_Lon': 'first'})
        if imd_temp_cols:
            agg.update({'IMD_TEMP_Ref_Lat': 'first',
                        'IMD_TEMP_Ref_Lon': 'first'})
        for col in other_cols:
            agg[f'{col}_Ref_Lat'] = 'first'
            agg[f'{col}_Ref_Lon'] = 'first'

        Data_pivots = (
            Data_pivots
            .groupby(id_cols, as_index=False)
            .agg(agg)
        )

        cols_order = id_cols + (
            ['ERA5_Ref_Lat','ERA5_Ref_Lon']
            if era5_cols else []
        ) + (
            ['ERA5L_Ref_Lat','ERA5L_Ref_Lon']
            if era5L_cols else []
        ) + (
            ['IMD_RAIN_Ref_Lat','IMD_RAIN_Ref_Lon']
            if imd_rain_col else []
        ) + (
            ['IMD_TEMP_Ref_Lat','IMD_TEMP_Ref_Lon']
            if imd_temp_cols else []
        ) + sum([[c, f'{c}_Ref_Lat', f'{c}_Ref_Lon'] for c in other_cols], []
                ) + all_data
        safe_cols = [col for col in cols_order if col in Data_pivots.columns]
        Data_pivots = Data_pivots[safe_cols]

    if code_mode != "claims":
        if save_data == 1:
            # Data_pivots.to_excel(path + r"\2a. Full Data.xlsx")

            if loc_ref == "Name":
                if loc_type == "Block":
                    Data_pivots = Data_pivots.sort_values(by=['State', 'District', 'Block', 'Date']).reset_index(drop=True)
                    grouped = Data_pivots.groupby(['State', 'District'])
                elif loc_type == "Pincode":
                    Data_pivots = Data_pivots.sort_values(by=['Pincode', 'Date']).reset_index(drop=True)
                    grouped = Data_pivots.groupby(['Pincode'])
            elif loc_ref == "Coordinates":
                Data_pivots = Data_pivots.sort_values(by=['State', 'District', 'Date']).reset_index(drop=True)
                grouped = Data_pivots.groupby(['State', 'District'])

            print("Historical data prepared.")

            with pd.ExcelWriter(base_data_path, engine='xlsxwriter') as writer:
                for name, group in grouped:
                    if loc_ref == "Name":
                        if loc_type == "Block":
                            state, district = name
                            sheet_name = f"{state}_{district}".replace(" ", "")[:31]
                        elif loc_type == "Pincode":
                            pincode = name[0]
                            sheet_name = f"{pincode}".replace("(),", "")[:31]
                    elif loc_ref == "Coordinates":
                        state, district = name
                        sheet_name = f"{state}_{district}".replace(" ", "")[:31]

                    group.to_excel(writer, sheet_name=sheet_name, index=False)

                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]

                    # Freeze the header row
                    worksheet.freeze_panes(1, 0)

                    # Apply autofilter to all columns
                    max_row = len(group)  # Adding 1 to include header row
                    max_col = len(group.columns) - 1
                    worksheet.autofilter(0, 0, max_row, max_col)

                    # Set active cell to A2
                    worksheet.set_selection(1, 0, 1, 0)

            print("Historical data saved.")

        if save_data_dump == 1:
            data_dump_file_path = os.path.join(path, 'Data_Working', "Risk_Datas.parquet")
            os.makedirs(os.path.dirname(data_dump_file_path), exist_ok=True)
            Risk_Datas.to_parquet(data_dump_file_path)

            # Dictionary of lists with variable names as keys
            lists_dict = {
                "Pincodes": Pincodes,
                "Locations": Locations,
                "State_Names": State_Names,
                "Dist_Names": Dist_Names,
                "Block_Names": Block_Names,
            }

            # Save each list with its variable name as the filename
            for name, data_list in lists_dict.items():
                file_path = os.path.join(path, 'Data_Working', f"{name}.pkl")
                file = open(file_path, 'wb')
                pickle.dump(data_list, file)
                file.close()

    data_end_time = datetime.now()
    data_hours = math.trunc((data_end_time - data_start_time).total_seconds()/3600)
    data_minutes = math.trunc((data_end_time - data_start_time).total_seconds()/60) - data_hours * 60
    data_seconds = math.trunc((data_end_time - data_start_time).total_seconds()%60)
    print("Data Phase Completed in: ", data_hours, "hours", data_minutes, "minutes", data_seconds, "seconds")
    ### DATA PHASE CONCLUDES ###

if run_analytics == 1:
    ### PRODUCT PHASE COMMENCES ####
    products_start_time = datetime.now()

    if run_data == 0:
        Risk_Datas = pd.read_parquet(path + "/Data_Working/Risk_Datas.parquet", engine='pyarrow')
        Risk_Datas = Risk_Datas.sort_index()
        # Dictionary to store loaded lists
        lists_dict = {
            "Pincodes": None,
            "Locations": None,
            "State_Names": None,
            "Dist_Names": None,
            "Block_Names": None,
        }

        # Load each list from its file in Data_Working
        for name in lists_dict.keys():
            file_path = os.path.join(path, 'Data_Working', f"{name}.pkl")
            with open(file_path, 'rb') as file:
                lists_dict[name] = pickle.load(file)

        # Assign the lists to individual variables
        State_Names = lists_dict["State_Names"]
        Dist_Names = lists_dict["Dist_Names"]
        Block_Names = lists_dict["Block_Names"]
        Pincodes = lists_dict["Pincodes"]
        Locations = lists_dict["Locations"]
        pre_data_end_time = datetime.now()
        pre_data_seconds = round((pre_data_end_time - products_start_time).total_seconds(), 2)
        print("Pre-saved Data Variables Loaded in", pre_data_seconds, "seconds.")
        
    os.makedirs(temp_path, exist_ok=True)


    for prod_index, prod_row in products.iterrows():

        # products = pd.read_csv(prod_file)
        # prod_index = 0
        # prod_row = products.loc[prod_index]

        if prod_row["to_run"] == 0:
            continue 

        termsheets_dump = []
        Backtest_dfs = []

        ## Pricing Model Selection ##
        code_mode = prod_row["code_mode"] #'reverse', 'custom', 'claims'

        ## User Inputs ##
        project_name = prod_row["project_name"]
        Product_ID = str(project_name) + "_" + str(prod_index+1)
        Risk_Name = prod_row["risk"]

        ## Pricing Assumptions ##
        TLR = float(prod_row["TLR"])
        if code_mode == "reverse":
            min_return_period = float(prod_row["min_return_period"]) if prod_row["min_return_period"] != "NA" else 1000

        RSD = pd.to_datetime(prod_row["RSD"], format="%d-%m-%Y")
        RED = pd.to_datetime(prod_row["RED"], format="%d-%m-%Y")

        ## Historical Analysis Assumptions ##
        if code_mode == 'claims':
            start_year = RSD.year
            end_year = RED.year
            num_years = 1
        elif code_mode in ('reverse', 'custom'):
            start_year = int(prod_row["start_year"])
            end_year = int(prod_row["end_year"])
            num_years = end_year - start_year + 1

        # loc_type = prod_row["loc_type"] # 'Pincode' or 'Block'
        # loc_ref = prod_row["loc_ref"] # "Name" or "Coordinates"

        Total_SI = float(prod_row["Total_SI"])

        Num_Risk_Units = int(prod_row["Num_Risk_Units"])
        SI_Weights = [float(w) for w in prod_row["SI_Weights"].strip('[]').split(';')]

        if code_mode == "reverse":
            Target_PR = float(prod_row["Target_PR"])
            Total_Target_BC = Total_SI * Target_PR * TLR
            Risk_Weights = [float(w) for w in prod_row["Risk_Weights"].strip('[]').split(';')]
            Risk_Weights = [w / sum(Risk_Weights) for w in Risk_Weights]
            Prod_Cust_Strikes = "["
            Prod_Cust_Payout_Weights = "["
            Prod_Xval = "["
        elif code_mode in ("custom", "claims"):
            if prod_row["Target_PR"] == "NA":
                Target_PR = np.nan
            else:
                Target_PR = float(prod_row["Target_PR"])
            
            Total_Target_BC = np.nan


        #### Risk Unit-Wise Calculation Starts #####
        for RU_index in range(Num_Risk_Units):
            # RU_index = 0

            Cover_ID = Product_ID + "." + str(RU_index + 1)

            cover_ts = pd.DataFrame(columns =  [
                "Cover_ID", "Cover", "Data_Source", "Ref_Lat", "Ref_Lon", "Xval", "Cover_Max_Payout", "Strike_1", "Payout_1", 
                "Strike_2", "Payout_2", "Strike_3", "Payout_3", "Historical_BC", "Modelled_BC"
            ])

            cover_backtest_payout = pd.DataFrame()
            cover_past_sheet = pd.DataFrame()

            bin_multipliers = [float(multiplier) for multiplier in prod_row["bin_multipliers"].strip('[]').split(';')[RU_index].split('&')] #Bin 1, 2, n-1, n
            strat_wt = int(prod_row["strat_wt"].strip('[]').split(';')[RU_index])

            RU_RSD, RU_RED = [pd.to_datetime(w, format="%d-%m-%Y") for w in prod_row["Phase_Dates"].strip('[]').split(';')[RU_index].split('&')]

            RU_SI = SI_Weights[RU_index] * Total_SI

            if code_mode == "reverse":
                RU_BC_Target = Total_Target_BC * Risk_Weights[RU_index]

                if len(termsheets_dump) > 0:
                    Risk_Wt_Elapsed = sum(Risk_Weights[0:RU_index])

                    RU_BC_Remaining = ((Total_Target_BC * Risk_Wt_Elapsed) - 
                                        sum(item['Modelled_BC'].item() for item in termsheets_dump if item['Product_ID'].item() == Product_ID))
                    RU_Risk_Wt_Remaining = 1 - Risk_Wt_Elapsed
                else:
                    RU_BC_Remaining = 0
                    RU_Risk_Wt_Remaining = Risk_Weights[RU_index]

                Cover_BC_Target = RU_BC_Target + (RU_BC_Remaining * (Risk_Weights[RU_index] / RU_Risk_Wt_Remaining))
                Cover_BC_Target = max(Cover_BC_Target, RU_BC_Target * 0.5)
                Cover_PR = (Cover_BC_Target/RU_SI)/TLR
            elif code_mode in ("custom", "claims"):
                Cover_BC_Target = Cover_PR = np.nan
                
            RU_Cover = prod_row["Covers"].strip('[]').split(';')[RU_index]
            RU_Data_Source = prod_row["Data_Sources"].strip('[]').split(';')[RU_index]
            RU_Loc_DF = Risk_Datas.loc[(Locations[f"{prod_index}.{RU_index}"][0], Locations[f"{prod_index}.{RU_index}"][1], RU_Data_Source)].reset_index()
            
            RU_DF = []

            for yr in range(start_year, end_year+1):
                # yr = 1974
                start_date = RU_RSD.replace(year=yr)

                day_adjust = 1 if (RU_RED.month == 2 and RU_RED.day == 29) else 0
                year_adjust = RU_RED.year - RU_RSD.year
                end_date = (RU_RED - timedelta(days=day_adjust)).replace(year=(yr + year_adjust))
                
                yr_filtered_RU_data = RU_Loc_DF[(RU_Loc_DF['time'] >= start_date) & (RU_Loc_DF['time'] <= end_date)]
                RU_DF.append(yr_filtered_RU_data)

            RU_DF = pd.concat(RU_DF, ignore_index=True)

            strike_1 = strike_2 = strike_3 = payout_1 = payout_2 = payout_3 = np.nan
            cover_threshold = 0
            if prod_row["Xval"].strip('[]').split(';')[RU_index] == 'NA':
                xval_orig = []  
            else: 
                xval_orig = [(RU_Cover, prod_row["Xval"].strip('[]').split(';')[RU_index])]
            xval = xval_orig.copy()
            if len(RU_DF) == 0:
                continue
            
            #### Cover Calculations Begin ####

            if code_mode == "reverse":

                if prod_row["Cover_threshold_min"].strip('[]').split(';')[RU_index] == "NA":
                    cover_threshold_min = -99 if RU_Data_Source in temp_ds_names else 0
                else:
                    cover_threshold_min = float(prod_row["Cover_threshold_min"].strip('[]').split(';')[RU_index])
                    
                if prod_row["Cover_threshold_max"].strip('[]').split(';')[RU_index] == "NA":
                    cover_threshold_max = 99 if RU_Data_Source in temp_ds_names else 9999
                else:
                    cover_threshold_max = float(prod_row["Cover_threshold_max"].strip('[]').split(';')[RU_index])

                if "VAR" in RU_Cover:

                    deductible = prod_row["Cover_deductible"].strip('[]').split(';')[RU_index]
                    num_levels = prod_row["Cover_levels"].strip('[]').split(';')[RU_index]
                    per_level_size = prod_row["Cover_level_size"].strip('[]').split(';')[RU_index]
                    
                    if "%" not in str(deductible):
                        deductible = int(deductible)

                        if "%" not in str(num_levels):
                            num_levels = int(num_levels)

                        if "%" not in str(per_level_size):
                        # Convert to float first to handle both '5' and '0.2'
                            per_level_size = float(per_level_size)
                            
                            # Optional: Convert back to int if it's a whole number (e.g., 5.0 -> 5)
                            if per_level_size.is_integer():
                                per_level_size = int(per_level_size)
                    ## E.g. if cover is required for 4, 8, 12 days, then 
                    # per_level_size = 4; Num_levels = 3; Deductible = 0 (in proportion of per level size)

                    if RU_Cover in ["GTE-HT_NHDVAR", "GTE-LT_NCDVAR", "GTE-ERH_NHRHDVAR", "GTE-HT_CHDVAR", "GTE-DRH_NLRHDVAR", "GTE-ER_CWDVAR",
                                    "GTE-ERH_CHRHDVAR", "GTE-ER_NWDVAR", "GTE-HT_NAHTVAR", "GTE-HT_PAHTVAR", "GTE-HT_PDHTVAR","GTE-HT_CPHTVAR", "GTE-ER_NWXDVAR", "GTE-ER_NWXDCONTVAR", 
                                    "GTE-HT_NHXDVAR"]:
                        levels_list = [deductible, num_levels, per_level_size]

                    base_payout = float(prod_row["Cover_basepay"].strip('[]').split(';')[RU_index]) * RU_SI
                    # base_pay to only be used when there is actually some flat base payout as soon as first trigger is breached 
                    # apart from that trigger's own per_level_payout

                    if ("%" not in str(deductible)) and ("%" not in str(num_levels)):
                        per_level_payout = round((RU_SI - base_payout) / num_levels, 3)
                    else:
                        per_level_payout = 0

                    discrete_payout = 1 if prod_row["Cover_payout_type"].strip('[]').split(';')[RU_index] == "Discrete" else 0
                    si_list = [base_payout, per_level_payout, discrete_payout]

                    if RU_Cover == "GTE-ER_NCRFVAR":
                        rolling_sum = RU_DF.groupby('year')['parameter'].rolling(window = int(xval_orig[0][1])).sum()
                        RU_cover_data = rolling_sum.groupby('year').max()
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, (100 - min(Cover_PR*500, 100))), -1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover == "GTE-HT_PAHTVAR":
                        # Calculate yearly average for the whole period
                        RU_cover_data = RU_DF.groupby('year')['parameter'].mean()
                        # Find a starting threshold based on the Target Premium Rate
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, (100 - min(Cover_PR*500, 100))), 1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover == "GTE-HT_PDHTVAR":
                        # Get the annual cumulative deviation data
                        RU_cover_data_df, _, _ = cover_definitions.HT_PDHTVAR(RU_DF, None)
                        RU_cover_data = RU_cover_data_df['parameter']
                        # Initial guess based on target premium percentile
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, (100 - min(Cover_PR*500, 100))), 1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover in ("LTE-DR_PCRFVAR", "LTE-LS_PCSRVAR"):
                        RU_cover_data = RU_DF.groupby('year')['parameter'].sum()
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, min(Cover_PR*500, 100)), 0), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover == "GTE-ER_PCRFVAR":
                        RU_cover_data = RU_DF.groupby('year')['parameter'].sum()
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, (100 - min(Cover_PR*500, 100))), 0), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover in ("GTE-HT_NHDVAR", "GTE-HT_NHXDVAR", "GTE-HT_CHDVAR", "GTE-ERH_NHRHDVAR", "GTE-ERH_CHRHDVAR", "GTE-DRH_NLRHDVAR", "GTE-HT_CPHTVAR"):
                        RU_cover_data = RU_DF['parameter']
                        cover_threshold = min(max(round(np.percentile(RU_DF["parameter"], (100 - min(Cover_PR*500, 100))), 1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover in ("GTE-ER_NWXDVAR", "GTE-ER_NWXDCONTVAR"):
                        rolling_sum = RU_DF.groupby('year')['parameter'].rolling(window = int(str(xval_orig[0][1]).strip('()').split(',')[0])).sum()
                        RU_cover_data = rolling_sum.groupby('year').max()
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, (100 - min(Cover_PR*500, 100))), -1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover in ("GTE-ER_NWDVAR", "GTE-ER_CWDVAR"):
                        RU_cover_data = RU_DF['parameter']
                        cover_threshold = min(max(round(np.percentile(RU_DF["parameter"], (100 - min(Cover_PR*500, 100))), -1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover in ("GTE-LT_NCDVAR", "GTE-LT_CCDVAR"):
                        RU_cover_data = RU_DF['parameter']
                        cover_threshold = min(max(round(np.percentile(RU_DF["parameter"], min(Cover_PR*500, 100)), 1), cover_threshold_min), cover_threshold_max)
                    elif RU_Cover == "LTE-LSM_NASMVAR":
                        rolling_sum = RU_DF.groupby('year')['parameter'].rolling(window = int(xval_orig[0][1])).sum()*100
                        RU_cover_data = rolling_sum.groupby('year').min()
                        cover_threshold = min(max(round(np.percentile(RU_cover_data, min(Cover_PR*500, 100)), -1), cover_threshold_min), cover_threshold_max)

                    iter = 0

                    while ((iter == 0) or (cover_threshold >= max((RU_cover_data.min() * bin_multipliers[0]), cover_threshold_min) and (cover_threshold <= 
                                                                        min((RU_cover_data.max() * bin_multipliers[3]), cover_threshold_max)))):

                        if "LTE" in RU_Cover and cover_threshold <= 0: 
                            break

                        if RU_Cover in ("LTE-DR_PCRFVAR", "GTE-ER_NCRFVAR", "LTE-LSM_NASMVAR", "LTE-LS_PCSRVAR", "GTE-HT_PAHTVAR", "GTE-HT_PDHTVAR", "GTE-ER_PCRFVAR"):
                            if "%" in str(num_levels):
                                num_levels_2 = round(cover_threshold * float(num_levels.strip('%')) / 100, 0)
                            else:
                                num_levels_2 = int(num_levels)

                            if "%" in str(per_level_size):
                                per_level_size_2 = round(cover_threshold * float(per_level_size.strip('%')) / 100, 0)
                            else:
                                per_level_size_2 = float(per_level_size)
    
                                if per_level_size_2.is_integer():
                                    per_level_size_2 = int(per_level_size_2)

                            per_level_payout = round((RU_SI - base_payout) / num_levels_2, 3)
                                
                            levels_list = [cover_threshold, num_levels_2, per_level_size_2]
                            si_list = [base_payout, per_level_payout, discrete_payout]

                        elif RU_Cover in ("GTE-HT_NHDVAR", "GTE-HT_CHDVAR", "GTE-ERH_NHRHDVAR", "GTE-ERH_CHRHDVAR", 
                                            "GTE-LT_NCDVAR", "GTE-LT_CCDVAR", "GTE-DRH_NLRHDVAR", "GTE-ER_NWDVAR", "GTE-ER_CWDVAR", "GTE-HT_CPHTVAR"):
                            xval =[(RU_Cover, cover_threshold)]

                        elif RU_Cover in ("GTE-HT_NHXDVAR", "GTE-ER_NWXDVAR"):
                            xval =[(RU_Cover, (int(str(xval_orig[0][1]).strip('()').split(',')[0]), float(cover_threshold)))]

                        elif RU_Cover in ("GTE-HT_NHXDVAR", "GTE-ER_NWXDCONTVAR"):
                            xval =[(RU_Cover, (int(str(xval_orig[0][1]).strip('()').split(',')[0]), str(xval_orig[0][1]).strip('()').split(',')[1],
                                    float(cover_threshold)))]

                        risk_defn = pd.DataFrame([
                            ("RU_"+str((RU_index+1)), (RU_RSD, RU_RED), RU_Cover, RU_Data_Source, levels_list, si_list)
                        ], columns=["Phase", "Dates", "Cover", "Data_Source", "Strikes", "Payouts"])
                        
                        RP_Dist, Payout_Freq, RU_cover_full = cover_rev_defn(RU_DF, risk_defn, xval, bin_multipliers, strat_wt)

                        if RP_Dist['Total_RP'][0] <= Cover_BC_Target: 
                            break
          
                        if (iter > 0) and ((cover_threshold == max((RU_cover_data.min() * bin_multipliers[0]), cover_threshold_min)) or 
                            (cover_threshold == min((RU_cover_data.max() * bin_multipliers[3]), cover_threshold_max))):
                            break
                        
                        if RU_Cover in ("GTE-ER_NCRFVAR", "GTE-ER_PCRFVAR"):
                            if cover_threshold <= 10:
                                cover_threshold = round(cover_threshold + 1, 0)
                            elif cover_threshold <= 30:
                                cover_threshold = round(cover_threshold + 2, 0)
                            elif cover_threshold <= 60:
                                cover_threshold = round(cover_threshold + 5, 0)
                            else:
                                cover_threshold = round(cover_threshold + 10, 0)
                        elif RU_Cover in ("LTE-DR_PCRFVAR", "LTE-LS_PCSRVAR"):
                            if cover_threshold <= 10:
                                cover_threshold = round(cover_threshold - 1, 0)
                            elif cover_threshold <= 30:
                                cover_threshold = round(cover_threshold - 2, 0)
                            elif cover_threshold <= 60:
                                cover_threshold = round(cover_threshold - 5, 0)
                            else:
                                cover_threshold = round(cover_threshold - 10, 0)
                        elif RU_Cover in ("GTE-HT_NHDVAR", "GTE-HT_CHDVAR","GTE-HT_PAHTVAR", "GTE-HT_CPHTVAR"):
                            cover_threshold = round(cover_threshold + 0.1, 1)
                        elif RU_Cover in ("GTE-HT_PDHTVAR"):
                            cover_threshold = round(cover_threshold + 1, 0)
                        elif RU_Cover in ("GTE-LT_NCDVAR", "GTE-LT_CCDVAR"):
                            cover_threshold = round(cover_threshold - 0.1, 1)                    
                        elif RU_Cover in ("GTE-ERH_NHRHDVAR", "GTE-ERH_CHRHDVAR"):
                            cover_threshold = round(cover_threshold + 1, 0)
                        elif RU_Cover in ("GTE-DRH_NLRHDVAR"):
                            cover_threshold = round(cover_threshold - 1, 0)
                        elif RU_Cover in ("GTE-ER_NWDVAR","GTE-ER_CWDVAR","GTE-ER_NWXDVAR","GTE-ER_NWXDCONTVAR"):
                            cover_threshold = round(cover_threshold + 10, 0)
                        elif "GTE" in RU_Cover:
                            cover_threshold = round(cover_threshold + 1, 0)
                        elif "LTE" in RU_Cover:
                            cover_threshold = round(cover_threshold - 1, 0)
                        
                        iter += 1

                    if RU_Cover in ("LTE-DR_PCRFVAR", "GTE-ER_NCRFVAR", "LTE-LSM_NASMVAR", "LTE-LS_PCSRVAR", "GTE-HT_PAHTVAR", "GTE-HT_PDHTVAR"):
                        strike_1 = cover_threshold
                    else:
                        strike_1 = deductible
                    if RU_Cover in ("LTE-DR_PCRFVAR", "GTE-ER_PCRFVAR", "GTE-ER_NCRFVAR", "LTE-LSM_NASMVAR", "LTE-LS_PCSRVAR"):
                        num_levels = num_levels_2
                        per_level_size = per_level_size_2
                    strike_2 = num_levels
                    strike_3 = per_level_size
                    payout_1 = base_payout
                    payout_2 = per_level_payout

                    risk_defn['DIST_RP'] = RP_Dist['Total_RP'].sum()
                    risk_defn['DIST_BC'] = Cover_BC_Target
                    Term_Sheet = risk_defn

                else:
                    Strike_SI_Weights = prod_row["Strike_SI_Weights"].strip('[]').split(';')[RU_index].split('&')
                    Strike_SI_Weights = [float(weight) for weight in Strike_SI_Weights]
                    Strike_SI_Weights = [w / max(Strike_SI_Weights) for w in Strike_SI_Weights]

                    Strike_Risk_Weights = prod_row["Strike_Risk_Weights"].strip('[]').split(';')[RU_index].split('&')
                    Strike_Risk_Weights = [float(weight) for weight in Strike_Risk_Weights]
                    Strike_Risk_Weights = [w / sum(Strike_Risk_Weights) for w in Strike_Risk_Weights]

                    cover_wt = pd.DataFrame(columns = ["Products", "Phase", "Cover", "Priority", "SI_Wt", "Risk_Wt", 
                                        "Data_Source", "Total_Phases", "N_Strikes"])

                    for i in reversed(range(len(Strike_SI_Weights))):
                        cover_wt.loc[len(cover_wt)] = {
                            "Products": Product_ID,
                            "Phase": Cover_ID,
                            "Cover": RU_Cover,
                            "Priority": str(len(Strike_SI_Weights)-i),
                            "SI_Wt": Strike_SI_Weights[i],
                            "Risk_Wt": Strike_Risk_Weights[i],
                            "Data_Source": RU_Data_Source,
                            "Total_Phases": 1,
                            "N_Strikes": 'S'+str((i+1))
                        }

                    risk_defn = pd.DataFrame([
                        (Product_ID, "RU_"+str((RU_index+1)), RU_RSD.strftime("%d-%m-%Y"), RU_RED.strftime("%d-%m-%Y"), 
                                (RU_RSD, RU_RED), RU_SI, Cover_PR)
                        ], columns=["Product_Name", "Phase", "RSD", "RED", "Dates", "Sum_Insured", "Premium_Rate"])

                    if RU_Cover in ("GTE-HT_NTHPM", "LTE-LT_NTHPM"):
                        Term_Sheet, RU_cover_full = cover_defn(RU_DF, risk_defn, cover_wt, TLR, xval, bin_multipliers, strat_wt,
                                                                min_return_period, cover_threshold_min, cover_threshold_max, payout_type = "Multiple")
                    else:
                        Term_Sheet, RU_cover_full = cover_defn(RU_DF, risk_defn, cover_wt, TLR, xval, bin_multipliers, strat_wt,
                                                                min_return_period, cover_threshold_min, cover_threshold_max)

                    strike_1 = Term_Sheet["Strikes"][0][0] if len(Term_Sheet["Strikes"][0]) >= 1 else np.nan
                    strike_2 = Term_Sheet["Strikes"][0][1] if len(Term_Sheet["Strikes"][0]) >= 2 else np.nan
                    strike_3 = Term_Sheet["Strikes"][0][2] if len(Term_Sheet["Strikes"][0]) >= 3 else np.nan
                    payout_1 = Term_Sheet["Payouts"][0][0] if len(Term_Sheet["Payouts"][0]) >= 1 else np.nan
                    payout_2 = Term_Sheet["Payouts"][0][1] if len(Term_Sheet["Payouts"][0]) >= 2 else np.nan
                    payout_3 = Term_Sheet["Payouts"][0][2] if len(Term_Sheet["Payouts"][0]) >= 3 else np.nan

                Prod_Cust_Strikes += str(strike_1)
                if not pd.isna(strike_2):
                    Prod_Cust_Strikes += "&" + str(strike_2) 
                if not pd.isna(strike_3):
                    Prod_Cust_Strikes += "&" + str(strike_3)
                Prod_Cust_Strikes += ";"

                if not pd.isna(payout_1):
                    Prod_Cust_Payout_Weights += str(round(payout_1/RU_SI, 8)) 
                if not pd.isna(payout_2):
                    Prod_Cust_Payout_Weights += "&" + str(round(payout_2/RU_SI, 8)) 
                if not pd.isna(payout_3):
                    Prod_Cust_Payout_Weights += "&" + str(round(payout_3/RU_SI, 8))
                Prod_Cust_Payout_Weights += ";"

                Prod_Xval += (str(xval[0][1]) if len(xval) > 0 else "NA") + ";"

            elif code_mode == "custom":

                # products = pd.read_csv(prod_file)
                # prod_index = 8
                # prod_row = products.loc[prod_index]

                Strikes = [float(w) for w in prod_row["Cust_Strikes"].strip('[]').split(';')[RU_index].split('&')]
                Payouts = [float(w) * RU_SI for w in prod_row["Cust_Payout_Weights"].strip('[]').split(';')[RU_index].split('&')]
 
                if 'VAR' in RU_Cover:
                    if len(Payouts) > 2:
                        Payouts[2] = int(prod_row["Cust_Payout_Weights"].strip('[]').split(';')[RU_index].split('&')[2])
                    else:
                        Payouts.append(1)
                risk_defn = pd.DataFrame([
                    (Cover_ID, (RU_RSD.strftime("%d-%m-%Y"), RU_RED.strftime("%d-%m-%Y")), RU_Cover, RU_Data_Source, Strikes, Payouts),
                ], columns=["Phase", "Dates", "Cover", "Data_Source", "Strikes", "Payouts"])

                xval = xval_orig

                if RU_Cover in ("GTE-HT_NTHPM", "LTE-LT_NTHPM"):
                    RP_Dist, Payout_Freq, RU_cover_full = cover_rev_defn(RU_DF, risk_defn, xval, bin_multipliers, strat_wt, payout_type = "Multiple")
                else:
                    RP_Dist, Payout_Freq, RU_cover_full = cover_rev_defn(RU_DF, risk_defn, xval, bin_multipliers, strat_wt)

                DIST_BC = RP_Dist[['Phase', 'Cover', 'Total_RP']]
                DIST_BC = DIST_BC.rename(columns = {'Total_RP': 'DIST_RP'})
                Term_Sheet = pd.merge(risk_defn, DIST_BC, on=['Phase', 'Cover'])
                Term_Sheet['DIST_BC'] = Cover_BC_Target

                if 'VAR' in RU_Cover:
                    strike_1 = Strikes[0] #Deductible
                    strike_2 = Strikes[1] #Num_levels
                    strike_3 = Strikes[2] if len(Strikes) >= 3 else 1 #Level Size
                    payout_1 = Payouts[0] #BasePay
                    payout_2 = Payouts[1] #Per Level Payout
                    payout_3 = Payouts[2] #1 for Discrete and 0 for Continuous

                    ## E.g. if cover is required for 4, 8, 12 days, then 
                    # per_level_size = 4; Num_levels = 3; Deductible = 0 (in proportion of per level size); 

                    # base_pay to only be used when there is actually some flat base payout as soon as first trigger is breached 
                    # apart from that trigger's own per_level_payout

                else:
                    strike_1 = Strikes[0]
                    strike_2 = Strikes[1] if len(Strikes) >= 2 else np.nan
                    strike_3 = Strikes[2] if len(Strikes) >= 3 else np.nan
                    payout_1 = Payouts[0]
                    payout_2 = Payouts[1] if len(Payouts) >= 2 else np.nan
                    payout_3 = Payouts[2] if len(Payouts) >= 3 else np.nan


            elif code_mode == "claims":

                # products = pd.read_csv(prod_file)
                # prod_index = 8
                # prod_row = products.loc[prod_index]

                Strikes = [float(w) for w in prod_row["Cust_Strikes"].strip('[]').split(';')[RU_index].split('&')]
                Payouts = [float(w) * RU_SI for w in prod_row["Cust_Payout_Weights"].strip('[]').split(';')[RU_index].split('&')]

                if 'VAR' in RU_Cover:
                    if len(Payouts) > 2:
                        Payouts[2] = int(prod_row["Cust_Payout_Weights"].strip('[]').split(';')[RU_index].split('&')[2])
                    else:
                        Payouts.append(1)

                risk_defn = pd.DataFrame([
                    (Cover_ID, (RU_RSD.strftime("%d-%m-%Y"), RU_RED.strftime("%d-%m-%Y")), RU_Cover, RU_Data_Source, Strikes, Payouts),
                ], columns=["Phase", "Dates", "Cover", "Data_Source", "Strikes", "Payouts"])

                Term_Sheet = risk_defn
                Term_Sheet['DIST_BC'] = 0
                Term_Sheet['DIST_RP'] = 0


                cover = risk_defn['Cover'].unique()[0]
                cover_type = cover.split('-')[1]
                xval = xval_orig
                val = next(map(lambda item: item[1], filter(lambda item: item[0] == cover, xval)), None)
                cover_data, max_criteria, RU_cover_full = getattr(cover_definitions, cover_type)(RU_DF, val)

                if 'VAR' in RU_Cover:
                    strike_1 = Strikes[0] #Deductible
                    strike_2 = Strikes[1] #Num_levels
                    strike_3 = Strikes[2] if len(Strikes) >= 3 else 1 #Level Size
                    payout_1 = Payouts[0] #BasePay
                    payout_2 = Payouts[1] #Per Level Payout
                    payout_3 = Payouts[2] #1 for Discrete and 0 for Continuous

                    ## E.g. if cover is required for 4, 8, 12 days, then 
                    # per_level_size = 4; Num_levels = 3; Deductible = 0 (in proportion of per level size); 

                    # base_pay to only be used when there is actually some flat base payout as soon as first trigger is breached 
                    # apart from that trigger's own per_level_payout

                else:
                    strike_1 = Strikes[0]
                    strike_2 = Strikes[1] if len(Strikes) >= 2 else np.nan
                    strike_3 = Strikes[2] if len(Strikes) >= 3 else np.nan
                    payout_1 = Payouts[0]
                    payout_2 = Payouts[1] if len(Payouts) >= 2 else np.nan
                    payout_3 = Payouts[2] if len(Payouts) >= 3 else np.nan

            #### BackTest, Product Analytics, and Termsheet Output Begin ####

            Back_Test, Past_Sheet, Summary = backtest(RU_DF, Term_Sheet, xval)
            Back_Test['Cover_ID'] = Cover_ID
            cover_backtest_payout = pd.concat([cover_backtest_payout, Back_Test], ignore_index=True)

            Past_Sheet['Cover_ID'] = Cover_ID
            cover_past_sheet = pd.concat([cover_past_sheet, Past_Sheet], ignore_index=True)

            Modelled_BC = round(Term_Sheet['DIST_RP'].sum(), 2)
            cover_ts.loc[len(cover_ts)] = {
                "Cover_ID": Cover_ID,
                "Cover": RU_cover_full,
                "Data_Source": RU_Data_Source,
                "Ref_Lat": Locations[f"{prod_index}.{RU_index}"][0],
                "Ref_Lon": Locations[f"{prod_index}.{RU_index}"][1],
                "Xval": xval[0][1] if len(xval) > 0 else "NA",
                "Cover_Max_Payout": RU_SI,
                "Strike_1": strike_1,
                "Payout_1": payout_1,
                "Strike_2": strike_2,
                "Payout_2": payout_2,
                "Strike_3": strike_3,
                "Payout_3": payout_3,
                "Historical_BC": Summary['Backtest_Payout'].sum(),
                "Modelled_BC": Modelled_BC
            }
            
            if code_mode == "reverse":
                Effective_PR = (RU_BC_Target/RU_SI)/TLR
            elif code_mode in ("custom", "claims"):
                if Target_PR == "NA":
                    Effective_PR = (cover_ts['Modelled_BC'][0]/cover_ts['Cover_Max_Payout'][0]) / TLR
                else:
                    Effective_PR = Target_PR

            Backtest_df = cover_backtest_payout[['Cover_ID', 'year', 'parameter', 'Backtest_Payout', 'Cover']]
            Backtest_df = Backtest_df.rename(columns={'parameter': 'observed_index'})
            Backtest_df['Cover'] = RU_cover_full

            Backtest_df['Risk'] = Risk_Name
            Backtest_df['Cover_Start_Date'] = RU_RSD
            Backtest_df['Cover_End_Date'] = RU_RED
            Backtest_df['Product_ID'] = Product_ID
            Backtest_df["Cover_Max_Payout"] =  RU_SI
            Backtest_df['State'] = State_Names[prod_index].title() if isinstance(State_Names[prod_index], str) else 'NA'
            Backtest_df['District'] = Dist_Names[prod_index].title() if isinstance(Dist_Names[prod_index], str) else 'NA'
            Backtest_df['Block'] = Block_Names[prod_index].title() if isinstance(Block_Names[prod_index], str) else 'NA'
            Backtest_df['Pincode'] = Pincodes[prod_index]

            if code_mode == "claims":
                Backtest_df["Data_Source"] = RU_Data_Source
                Backtest_df["Xval"] = xval[0][1] if len(xval) > 0 else "NA",
                Backtest_df["Strike_1"] = strike_1
                Backtest_df["Payout_1"] = payout_1
                Backtest_df["Strike_2"] = strike_2
                Backtest_df["Payout_2"] = payout_2
                Backtest_df["Strike_3"] = strike_3
                Backtest_df["Payout_3"] = payout_3
                Backtest_df = Backtest_df.rename(columns={'Backtest_Payout':'Calculated_Payout'})


            if code_mode == "claims":
                Backtest_df = Backtest_df[['Product_ID', 'Cover_ID', 'State', 'District', 'Block', 'Pincode', 'Risk',
                                    'Cover_Start_Date', 'Cover_End_Date', 'Cover', 'Xval', 'Data_Source', 'Cover_Max_Payout', 'Strike_1', 'Payout_1', 
                                    'Strike_2', 'Payout_2', 'Strike_3', 'Payout_3', 'year', 'observed_index', 'Calculated_Payout']]
            elif code_mode in ("custom", "reverse"):
                Backtest_df = Backtest_df[['Product_ID', 'Cover_ID', 'State', 'District', 'Block', 'Pincode', 'Risk', 
                                            'Cover_Start_Date', 'Cover_End_Date', 'Cover', 'Cover_Max_Payout', 'year', 
                                            'observed_index', 'Backtest_Payout']]

            Backtest_dfs.append(Backtest_df)

            if len(cover_past_sheet) > 0:
                Past_Sheet = cover_past_sheet.merge(cover_backtest_payout[['Cover_ID', 'year', 'parameter']], on=['year', 'Cover_ID'], how='left')
                Past_selected = Past_Sheet[['Cover_ID', 'year', 'parameter', 'Backtest_Payout']]
                Past_selected = Past_selected.rename(columns={'Backtest_Payout': 'Payout', 'parameter': 'observed_index'}).reset_index(drop=True)
                Past_selected['Payout'] = round(Past_selected['Payout'], 2)
                Past_selected = Past_selected.groupby('Cover_ID').apply(lambda x: x.drop('Cover_ID', axis=1).to_string(index=False)).reset_index(name='Historical_Payout_Details')
            else:
                Past_selected = pd.DataFrame({'Cover_ID': [Cover_ID], 'Historical_Payout_Details': ['NA']})

            #### Termsheet Output Data Begins ####
            ts_data = cover_ts
            
            ts_data['Product_ID'] = Product_ID
            ts_data['State'] = State_Names[prod_index].title() if isinstance(State_Names[prod_index], str) else 'NA'
            ts_data['District'] = Dist_Names[prod_index].title() if isinstance(Dist_Names[prod_index], str) else 'NA'
            ts_data['Block'] = Block_Names[prod_index].title() if isinstance(Block_Names[prod_index], str) else 'NA'
            ts_data['Pincode'] = Pincodes[prod_index]
            ts_data['Risk'] = Risk_Name
            ts_data['Risk_Start_Date'] = RSD
            ts_data['Risk_End_Date'] = RED
            ts_data['Total_Sum_Insured'] = Total_SI
            ts_data['Cover_Net_Premium'] = round(ts_data['Cover_Max_Payout'] * Effective_PR, 2)
            ts_data['Cover_Gross_Premium'] = round(ts_data['Cover_Net_Premium'] * 1.18, 2)
            ts_data['Cover_Start_Date'] = RU_RSD
            ts_data['Cover_End_Date'] = RU_RED
            ts_data = ts_data.merge(Past_selected, on=['Cover_ID'], how='left')

            if len(cover_past_sheet) > 0:
                ts_data = ts_data.merge(Past_Sheet.groupby('Cover_ID')['Backtest_Payout'].max().reset_index(name='Max_BT_Payout'), on='Cover_ID', how='left')
                ts_data['ValueAtRisk'] = ts_data.apply(
                    lambda row: min(row['Max_BT_Payout'] * 1.2, row['Cover_Max_Payout']) if pd.notna(row['Max_BT_Payout']) 
                                else row['Cover_Max_Payout'] * 0.2, 
                    axis=1
                )

                ts_data = ts_data.merge(Past_Sheet.groupby('Cover_ID').size().reset_index(name='Num_Payout_Years'), on='Cover_ID', how='left')

                ts_data = ts_data.merge(Past_Sheet.groupby('Cover_ID')['Backtest_Payout'].sum().reset_index(name='Total_BT_Payout'), on='Cover_ID', how='left')
                ts_data['PayoutAvgSeverity'] = ts_data.apply(
                    lambda row: round(row['Total_BT_Payout'] / row['Num_Payout_Years'], 2) if row['Num_Payout_Years'] > 0 else 0,
                    axis=1
                )

                ts_data['YearsAboveAvgSeverity'] = ts_data.apply(
                    lambda row: (Past_Sheet[Past_Sheet['Cover_ID'] == row['Cover_ID']]['Backtest_Payout'] > row['PayoutAvgSeverity']).sum(),
                    axis=1
                )

                ts_data['PayoutFrequency'] = round(ts_data['Num_Payout_Years'] / num_years, 4)
                ts_data['Historical_BC_pct'] = round(ts_data['Historical_BC'] / RU_SI, 4)
                ts_data['ValueAtRisk_pct'] = round(ts_data['ValueAtRisk'] / RU_SI, 4)
                ts_data['PayoutAvgSeverity_pct'] = round(ts_data['PayoutAvgSeverity'] / RU_SI, 4)
                ts_data['YearsAboveAvgSeverity_pct'] = round(ts_data['YearsAboveAvgSeverity']/num_years, 4)
                ts_data['Modelled_BC_pct'] = round(ts_data['Modelled_BC'] / RU_SI, 4)
                ts_data['Modelled_BC_contrib_pct'] = round(ts_data['Modelled_BC'] / Total_SI, 4) 

            else:
                ts_data['Historical_BC_pct'] = 0
                ts_data['Num_Payout_Years'] = 0
                ts_data['PayoutFrequency'] = 0
                ts_data['PayoutAvgSeverity'] = 0
                ts_data['PayoutAvgSeverity_pct'] = 0
                ts_data['ValueAtRisk'] = RU_SI * 0.2
                ts_data['ValueAtRisk_pct'] = 0.2
                ts_data['YearsAboveAvgSeverity'] = 0
                ts_data['YearsAboveAvgSeverity_pct'] = 0
                ts_data['Modelled_BC_pct'] = round(ts_data['Modelled_BC'] / RU_SI, 4)
                ts_data['Modelled_BC_contrib_pct'] = round(ts_data['Modelled_BC'] / Total_SI, 4) 

            ts_order = ['Product_ID', 'Cover_ID', 'State', 'District', 'Block', 'Pincode', 'Ref_Lat', 'Ref_Lon', 'Risk', 
                        'Risk_Start_Date', 'Risk_End_Date', 'Total_Sum_Insured', 'Cover_Net_Premium', 'Cover_Gross_Premium', 
                        'Cover_Start_Date', 'Cover_End_Date', 'Cover', 'Data_Source', 'Xval', 'Cover_Max_Payout', 
                        'Strike_1', 'Payout_1', 'Strike_2', 'Payout_2', 'Strike_3', 'Payout_3',
                        'Historical_Payout_Details', 'Historical_BC', 'Historical_BC_pct', 'Num_Payout_Years', 'PayoutFrequency', 
                        'PayoutAvgSeverity', 'PayoutAvgSeverity_pct', 'ValueAtRisk', 'ValueAtRisk_pct', 'YearsAboveAvgSeverity', 
                        'YearsAboveAvgSeverity_pct', 'Modelled_BC', 'Modelled_BC_pct', 'Modelled_BC_contrib_pct']
            ts_data = ts_data[ts_order]

            termsheets_dump.append(ts_data)

            # print("Product ", prod_index+1, "Risk Unit", RU_index+1, "completed")
            RU_index += 1

        if code_mode == "reverse":

            Prod_Cust_Strikes = Prod_Cust_Strikes[:-1] + "]"
            products_2.loc[prod_index, 'Cust_Strikes'] = Prod_Cust_Strikes

            Prod_Cust_Payout_Weights = Prod_Cust_Payout_Weights[:-1] + "]"
            products_2.loc[prod_index, 'Cust_Payout_Weights'] = Prod_Cust_Payout_Weights

            Prod_Xval = Prod_Xval[:-1] + "]"
            products_2.loc[prod_index, 'Xval'] = str(Prod_Xval)

        if len(termsheets_dump) > 0:
            termsheets_dump = pd.concat(termsheets_dump, ignore_index=True)

            unique_ts_name = f"termsheets_dump_{prod_index}.parquet"
            termsheets_dump["Xval"] = pd.to_numeric(termsheets_dump["Xval"], errors="coerce")
            termsheets_dump.to_parquet(os.path.join(temp_path, unique_ts_name), index=False)
            termsheets_dump = []

            Backtest_dfs = pd.concat(Backtest_dfs, ignore_index=True)
            unique_bt_name = f"Backtest_dfs_{prod_index}.parquet"
            Backtest_dfs.to_parquet(os.path.join(temp_path, unique_bt_name), index=False)
            Backtest_dfs = []

        print("Product", (prod_index + 1), "/", products["to_run"].sum(), "processed.")

    code_mode = products['code_mode'].iloc[0]

    if code_mode != 'claims':
        products_2.fillna('NA', inplace=True)
        products_2.to_csv(prod_file_2, index=False)

    termsheets_dump = []
    Backtest_dfs = []

    files = [f for f in os.listdir(temp_path) if f.endswith(".parquet")]
    sorted_files = sorted(files, key=lambda fname: int(fname.split('_')[2].split('.')[0]))

    for fname in sorted_files: 
        file_path = os.path.join(temp_path, fname)

        if fname.startswith("termsheets_dump_"):
            termsheets_dump.append(pd.read_parquet(file_path))
        elif fname.startswith("Backtest_dfs_"):
            Backtest_dfs.append(pd.read_parquet(file_path))
            
    termsheets_dump = pd.concat(termsheets_dump, ignore_index=True)
    termsheets_dump.fillna('NA', inplace=True)

    Backtest_dfs = pd.concat(Backtest_dfs, ignore_index=True)
    Backtest_dfs.fillna('NA', inplace=True)

    ### PRODUCT PHASE CONCLUDES ####




    ### TERMSHEET ANALYTICS COMMENCE ###

    if code_mode in ('reverse', 'custom'):
        ts_dump_order = ['Product_ID', 'State', 'District', 'Block', 'Pincode', 'Risk', 'Risk_Start_Date', 'Risk_End_Date', 
                         'Total_Sum_Insured']

        termsheets_analytics = termsheets_dump[ts_dump_order].drop_duplicates().reset_index(drop=True)

        backtest_yearly = Backtest_dfs.groupby(['Product_ID', 'year'])['Backtest_Payout'].sum().reset_index()

        backtest_analytics = backtest_yearly.groupby(['Product_ID']).agg(
            Sum_Historical_Payout = ('Backtest_Payout', 'sum'),
            Max_Historical_Payout = ('Backtest_Payout', 'max'),
            Num_Payout_Years = ('Backtest_Payout', lambda x: (x > 0).sum()),
            PayoutAvgSeverity = ('Backtest_Payout', lambda x: round(x[x > 0].mean(), 2))
        ).reset_index()
        numeric_columns = ['Sum_Historical_Payout', 'Max_Historical_Payout', 'Num_Payout_Years']
        backtest_analytics[numeric_columns] = backtest_analytics[numeric_columns].apply(pd.to_numeric, errors='coerce')

        backtest_yearly = backtest_yearly.merge(backtest_analytics[['Product_ID', 'PayoutAvgSeverity']], 
                                                on=['Product_ID'], how='left')

        backtest_analytics_2 = backtest_yearly.groupby(['Product_ID']).agg(
            YearsAboveAvgSeverity = ('Backtest_Payout', lambda x: (x > backtest_yearly.loc[x.index, 'PayoutAvgSeverity']).sum())
        ).reset_index()

        backtest_analytics = backtest_analytics.merge(backtest_analytics_2, on=['Product_ID'], how='left')
        backtest_analytics['Historical_BC'] = round(backtest_analytics['Sum_Historical_Payout']/num_years, 2)
        backtest_analytics['VAR'] = round(backtest_analytics['Max_Historical_Payout'] * 1.2, 2)

        backtest_analytics = backtest_analytics[['Product_ID', 'Historical_BC', 'VAR', 'Num_Payout_Years',
                                                'PayoutAvgSeverity', 'YearsAboveAvgSeverity']]

        backtest_analytics.fillna(0, inplace=True)

        termsheets_stats = termsheets_dump[['Product_ID', 'Modelled_BC', 'Cover_Net_Premium', 
                                            'Cover_Gross_Premium']].groupby(['Product_ID']).sum().reset_index()

        termsheets_stats = termsheets_stats.rename(columns = {'Cover_Net_Premium': 'Net_Premium', 
                                                            'Cover_Gross_Premium': 'Gross_Premium'})

        termsheets_analytics = pd.merge(termsheets_analytics, backtest_analytics, on=['Product_ID'], how='left')
        termsheets_analytics = pd.merge(termsheets_analytics, termsheets_stats, on=['Product_ID'], how='left')

        termsheets_analytics['ValueAtRisk'] = np.maximum(np.minimum(termsheets_analytics['VAR'], termsheets_analytics['Total_Sum_Insured']),
                                                        termsheets_analytics['Total_Sum_Insured'] * 0.2)

        termsheets_analytics['PayoutFrequency'] = round(termsheets_analytics['Num_Payout_Years']/num_years, 4)
        termsheets_analytics['Historical_BC_pct'] = round(termsheets_analytics['Historical_BC'] / termsheets_analytics['Total_Sum_Insured'], 4)
        termsheets_analytics['ValueAtRisk_pct'] = round(termsheets_analytics['ValueAtRisk'] / termsheets_analytics['Total_Sum_Insured'], 4)
        termsheets_analytics['PayoutAvgSeverity_pct'] = round(termsheets_analytics['PayoutAvgSeverity'] / termsheets_analytics['Total_Sum_Insured'], 4)
        termsheets_analytics['YearsAboveAvgSeverity_pct'] = round(termsheets_analytics['YearsAboveAvgSeverity']/num_years, 4)
        termsheets_analytics['Modelled_BC_pct'] = round(termsheets_analytics['Modelled_BC'] / termsheets_analytics['Total_Sum_Insured'], 4)

        termsheets_analytics = termsheets_analytics[ts_dump_order + ['Net_Premium', 'Gross_Premium', 'Historical_BC', 
                                                    'Historical_BC_pct', 'Num_Payout_Years', 'PayoutFrequency', 'PayoutAvgSeverity', 
                                                    'PayoutAvgSeverity_pct', 'ValueAtRisk', 'ValueAtRisk_pct', 'YearsAboveAvgSeverity', 
                                                    'YearsAboveAvgSeverity_pct', 'Modelled_BC', 'Modelled_BC_pct']]

        print("Product analytics prepared.")
    ### TERMSHEET ANALYTICS CONCLUDE ###



    ### SAVING ANALYTICS COMMENCES ###
    if 1 == 1:
        def auto_adjust_column_width(worksheet):
            """Auto-adjust column widths based on cell contents."""
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2  # Add some padding to the width
                worksheet.column_dimensions[column].width = adjusted_width

        if code_mode == 'claims':
            with pd.ExcelWriter(claims_path, engine='openpyxl') as writer:
                workbook = writer.book

                # Define number format
                date_format = NamedStyle(name="date_format", number_format='DD-MM-YYYY')
                workbook.add_named_style(date_format)

                number_format = NamedStyle(name="number_format", number_format='#,##0.00')
                workbook.add_named_style(number_format)

                percentage_format = NamedStyle(name="percentage_format", number_format='0.00%')
                workbook.add_named_style(percentage_format)



                ## Formatting Backtest Sheet ##
                Backtest_dfs.to_excel(writer, sheet_name='Working', index=False)
                worksheet = writer.sheets['Working']

                # Applying number formats

                format_start = 8

                for col in [0, 1]:
                    for cell in worksheet[get_column_letter(format_start + col)]:
                        cell.style = date_format

                for col in [5, 6, 7, 8, 9, 10, 11]:
                    for cell in worksheet[get_column_letter(format_start + col)]:
                        cell.style = number_format

                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')

                # Adjusting the view
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.freeze_panes = 'A2'
                for cell in worksheet[1]:
                    cell.border = None
                auto_adjust_column_width(worksheet)
                worksheet.sheet_view.selection[0].activeCell = "A2"
                worksheet.sheet_view.selection[0].sqref = "A2"


                Data_pivots.drop_duplicates(inplace=True)
                Data_pivots.to_excel(writer, sheet_name='Data', index=False)
                worksheet = writer.sheets['Data']

                # Adjusting the view
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.freeze_panes = 'A2'
                for cell in worksheet[1]:
                    cell.border = None
                auto_adjust_column_width(worksheet)
                worksheet.sheet_view.selection[0].activeCell = "A2"
                worksheet.sheet_view.selection[0].sqref = "A2"


        elif code_mode in ('reverse', 'custom'):
            if save_analytics == 1:
                with pd.ExcelWriter(analytics_path, engine='openpyxl') as writer:
                    workbook = writer.book

                    # Define number format
                    date_format = NamedStyle(name="date_format", number_format='DD-MM-YYYY')
                    workbook.add_named_style(date_format)

                    number_format = NamedStyle(name="number_format", number_format='#,##0.00')
                    workbook.add_named_style(number_format)

                    percentage_format = NamedStyle(name="percentage_format", number_format='0.00%')
                    workbook.add_named_style(percentage_format)

                    ## Formatting Termsheet Analytics Sheet ##
                    termsheets_analytics.to_excel(writer, sheet_name='Termsheet_Wise_Analytics', index=False)
                    worksheet = writer.sheets['Termsheet_Wise_Analytics']

                    # Applying number formats
                    for col in [7, 8]:
                        for cell in worksheet[get_column_letter(col)]:
                            cell.style = date_format

                    for col in [9, 10, 11, 12, 14, 16, 18, 20, 22]:  
                        for cell in worksheet[get_column_letter(col)]:
                            cell.style = number_format

                    for col in [13, 15, 17, 19, 21, 23]:
                        for cell in worksheet[get_column_letter(col)]:
                            cell.style = percentage_format

                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='left')

                    # Adjusting the view
                    worksheet.auto_filter.ref = worksheet.dimensions
                    worksheet.freeze_panes = 'A2'
                    for cell in worksheet[1]:
                        cell.border = None
                    auto_adjust_column_width(worksheet)
                    worksheet.sheet_view.selection[0].activeCell = "A2"
                    worksheet.sheet_view.selection[0].sqref = "A2"



                    ## Formatting Cover_Wise_Analytics (Termsheets Dump) Sheet ##
                    termsheets_dump.to_excel(writer, sheet_name='Cover_Wise_Analytics', index=False)
                    worksheet = writer.sheets['Cover_Wise_Analytics']

                    # Applying number formats
                    for col in [9, 10, 11, 15, 16]:
                        for cell in worksheet[get_column_letter(col)]:
                            cell.style = date_format

                    for col in [12, 13, 14, 20, 21, 22, 23, 24, 25, 26, 28, 30, 32, 34, 36, 38]:
                        for cell in worksheet[get_column_letter(col)]:
                            cell.style = number_format

                    for col in [29, 31, 33, 35, 37, 39, 40]:
                        for cell in worksheet[get_column_letter(col)]:
                            cell.style = percentage_format

                    for cell in worksheet[get_column_letter(18)]:
                        cell.alignment = Alignment(wrap_text=True)

                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='left')

                    for row in worksheet.iter_rows():
                        worksheet.row_dimensions[row[0].row].height = 14.4  # Adjust the height as needed

                    # Adjusting the view
                    worksheet.auto_filter.ref = worksheet.dimensions
                    worksheet.freeze_panes = 'A2'
                    for cell in worksheet[1]:
                        cell.border = None
                    auto_adjust_column_width(worksheet)
                    worksheet.column_dimensions[get_column_letter(27)].width = 32
                    worksheet.sheet_view.selection[0].activeCell = "A2"
                    worksheet.sheet_view.selection[0].sqref = "A2"



                    ## Formatting Backtest Sheet ##
                    if (((products["end_year"] - products["start_year"] + 1) * products["Num_Risk_Units"])[products["to_run"] == 1]).sum() <= 100000: 
                        Backtest_dfs.to_excel(writer, sheet_name='Backtest', index=False)
                        worksheet = writer.sheets['Backtest']

                        # Applying number formats
                        for col in [8, 9]:
                            for cell in worksheet[get_column_letter(col)]:
                                cell.style = date_format

                        for col in [11, 13, 14]:
                            for cell in worksheet[get_column_letter(col)]:
                                cell.style = number_format

                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='left')

                        # Adjusting the view
                        worksheet.auto_filter.ref = worksheet.dimensions
                        worksheet.freeze_panes = 'A2'
                        for cell in worksheet[1]:
                            cell.border = None
                        auto_adjust_column_width(worksheet)
                        worksheet.sheet_view.selection[0].activeCell = "A2"
                        worksheet.sheet_view.selection[0].sqref = "A2"
                if (((products["end_year"] - products["start_year"] + 1) * products["Num_Risk_Units"])[products["to_run"] == 1]).sum() > 100000: 
                    Backtest_dfs.to_parquet(path + r"\3b. Backtest.parquet")

                print("Product analytics saved.")

            elif save_analytics == 0:
                if loc_type == 'Pincode':
                    temp = termsheets_dump[['Pincode', 'Cover', 'Modelled_BC', 'Total_Sum_Insured', 'Cover_Max_Payout']]
                else:
                    temp = termsheets_dump[['State', 'District', 'Block', 'Cover', 'Modelled_BC', 'Total_Sum_Insured', 'Cover_Max_Payout']]

                prod_TSI = products['Total_SI'].sum()
                temp['Modelled_BC_pct'] = round(temp['Modelled_BC'] * 100 / temp['Cover_Max_Payout'], 2)
                temp['Modelled_BC_contrib_pct'] = round(temp['Modelled_BC'] * 100 / prod_TSI, 2)

                temp['Cover_sort'] = temp['Cover'].str.split('(').str[0]

                if loc_type == 'Pincode':
                    print(temp[['Pincode', 'Cover', 'Cover_sort', 'Modelled_BC_pct', 'Modelled_BC_contrib_pct']].sort_values(by=['Cover_sort', 'Pincode']).drop(columns='Cover_sort').reset_index(drop=True))
                else:
                    print(temp[['State', 'District', 'Block', 'Cover', 'Cover_sort', 'Modelled_BC_pct', 'Modelled_BC_contrib_pct']].sort_values(by=['Cover_sort', 'State', 'District', 'Block']).drop(columns='Cover_sort').reset_index(drop=True))

                print("Historical BC: ",round(termsheets_analytics['Historical_BC'].sum() * 100 / termsheets_analytics['Total_Sum_Insured'].sum(), 2),"%")
                print("Modelled BC: ",round(termsheets_analytics['Modelled_BC'].sum()*100 / termsheets_analytics['Total_Sum_Insured'].sum(), 2),"%")
                
    ### SAVING ANALYTICS CONCLUDES ###

    products_end_time = datetime.now()

    product_hours = math.trunc((products_end_time - products_start_time).total_seconds()/3600)
    product_minutes = math.trunc((products_end_time - products_start_time).total_seconds()/60) - product_hours * 60
    product_seconds = math.trunc((products_end_time - products_start_time).total_seconds()%60)


if run_data == 1:
    print("Data Phase Completed in: ", data_hours, "hours", data_minutes, "minutes", data_seconds, "seconds")
if run_analytics == 1:
    print("Product Phase Completed in: ", product_hours, "hours", product_minutes, "minutes", product_seconds, "seconds")

if run_data == 1:
    total_start_time = data_start_time
else:
    total_start_time = products_start_time

if run_analytics == 1:
    total_end_time = products_end_time
else:
    total_end_time = data_end_time

total_hours = math.trunc((total_end_time - total_start_time).total_seconds()/3600)
total_minutes = math.trunc((total_end_time - total_start_time).total_seconds()/60) - total_hours * 60
total_seconds = math.trunc((total_end_time - total_start_time).total_seconds()%60)

print("Total Runtime: ", total_hours, "hours", total_minutes, "minutes", total_seconds, "seconds")

gc.collect()
shutil.rmtree(temp_path, ignore_errors=True)

# Run output.py only if termsheet is "Yes"
if termsheet == "Yes":
    subprocess.run([sys.executable, "output.py", pdf])